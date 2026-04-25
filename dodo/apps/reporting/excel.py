"""
Excel report builder for the Reporting app.

Produces well-formatted, multi-sheet workbooks for each report type:
- progress      → Dashboard + Status Analysis + per-cluster Q1–Q4 tabs
- verification  → Dashboard + per-stage breakdown + per-cluster sheets
- indicators    → Dashboard + matrix sheets per tier
- donor         → Cover + per-donor schedule sheets

Mirrors the visual style of the CO_2026 Reporting Tracker template.
"""
from collections import defaultdict
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule


# ---------------------------------------------------------------------------
# Style constants — match the look of your existing CO tracker template
# ---------------------------------------------------------------------------

DODO_BLUE = '0077C8'
DODO_BLUE_LIGHT = 'E6F2FB'
HEADER_DARK = '1E293B'
GREY_BG = 'F8FAFC'
WHITE = 'FFFFFF'

STATUS_FILLS = {
    'submitted':       'DCFCE7',  # green-100
    'under_review':    'DBEAFE',  # blue-100
    'pending':         'FEF3C7',  # amber-100
    'overdue':         'FEE2E2',  # red-100
    'not_started':     'F1F5F9',  # slate-100
    'closed':          'E0E7FF',  # indigo-100
    'not_applicable':  'F3F4F6',  # gray-100
    'completed':       'DCFCE7',
    'field_verification':    'DBEAFE',
    'documentation_review':  'DBEAFE',
    'validation_meeting':    'DBEAFE',
}
STATUS_FONTS = {
    'submitted':       '166534',
    'under_review':    '1E40AF',
    'pending':         '854D0E',
    'overdue':         '991B1B',
    'not_started':     '64748B',
    'closed':          '3730A3',
    'not_applicable':  '6B7280',
    'completed':       '166534',
    'field_verification':    '1E40AF',
    'documentation_review':  '1E40AF',
    'validation_meeting':    '1E40AF',
}

THIN = Side(style='thin', color='E2E8F0')
ALL_BORDERS = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _title(ws, text, row=1, span=8):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name='Calibri', size=16, bold=True, color=WHITE)
    cell.fill = PatternFill('solid', start_color=DODO_BLUE)
    cell.alignment = Alignment(vertical='center', indent=1)
    ws.row_dimensions[row].height = 30
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def _subtitle(ws, text, row=2, span=8):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name='Calibri', size=10, italic=True, color='64748B')
    cell.alignment = Alignment(vertical='center', indent=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def _section(ws, text, row, span=8):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name='Calibri', size=12, bold=True, color=DODO_BLUE)
    cell.fill = PatternFill('solid', start_color=DODO_BLUE_LIGHT)
    cell.alignment = Alignment(vertical='center', indent=1)
    ws.row_dimensions[row].height = 22
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def _header_row(ws, headers, row, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = Font(name='Calibri', size=10, bold=True, color=WHITE)
        c.fill = PatternFill('solid', start_color=HEADER_DARK)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = ALL_BORDERS
    ws.row_dimensions[row].height = 30


def _body_cell(ws, value, row, col, *, bold=False, fill=None, font_color=None,
               align='left', wrap=False, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(
        name='Calibri', size=10, bold=bold,
        color=font_color or '0F172A',
    )
    if fill:
        c.fill = PatternFill('solid', start_color=fill)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    c.border = ALL_BORDERS
    if number_format:
        c.number_format = number_format
    return c


def _kpi_tile(ws, row, col, label, value, color=DODO_BLUE):
    """A small 2-row KPI 'tile' starting at (row, col). Spans 2 cols."""
    lbl = ws.cell(row=row, column=col, value=label.upper())
    lbl.font = Font(name='Calibri', size=8, bold=True, color='64748B')
    lbl.alignment = Alignment(horizontal='center', vertical='center')
    lbl.fill = PatternFill('solid', start_color=GREY_BG)
    lbl.border = ALL_BORDERS

    val = ws.cell(row=row + 1, column=col, value=value)
    val.font = Font(name='Calibri', size=16, bold=True, color=color)
    val.alignment = Alignment(horizontal='center', vertical='center')
    val.fill = PatternFill('solid', start_color=WHITE)
    val.border = ALL_BORDERS

    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
    ws.row_dimensions[row].height = 18
    ws.row_dimensions[row + 1].height = 32


def _status_pill(ws, row, col, status_code, status_label):
    fill = STATUS_FILLS.get(status_code, 'F1F5F9')
    color = STATUS_FONTS.get(status_code, '64748B')
    return _body_cell(ws, status_label or '—', row, col,
                      bold=True, fill=fill, font_color=color, align='center')


def _set_widths(ws, widths):
    """widths: dict of {column_letter_or_index: width}."""
    for k, w in widths.items():
        letter = k if isinstance(k, str) else get_column_letter(k)
        ws.column_dimensions[letter].width = w


def _freeze_header(ws, row=5, col=2):
    ws.freeze_panes = ws.cell(row=row, column=col).coordinate


def _safe_sheet_name(name):
    """Excel sheet names: ≤31 chars, no : \\ / ? * [ ]"""
    bad = set(':\\/?*[]')
    cleaned = ''.join('_' if ch in bad else ch for ch in str(name))
    return cleaned[:31] or 'Sheet'


def _val(obj, attr, default=''):
    """Defensive attribute getter — survives None and missing attrs."""
    if obj is None:
        return default
    v = getattr(obj, attr, default)
    return v if v is not None else default


def _wb_to_bytes(wb):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ===========================================================================
# Progress report
# ===========================================================================

PROGRESS_HEADERS = [
    'PIMS ID', 'Project', 'Programme Unit', 'Donor',
    'Project Status', 'Report Status', 'Last Updated', 'Updated By', 'Notes',
]


def build_progress_workbook(*, country_office, year, quarter,
                             projects, statuses_by_project, cycle):
    """
    projects: iterable of Project (already filtered to CO + ordered by unit)
    statuses_by_project: {project_id: ProjectReportingStatus} for the cycle
    cycle: ReportingCycle or None
    """
    wb = Workbook()

    # ---- Dashboard ----
    ws = wb.active
    ws.title = 'Dashboard'
    co_name = _val(country_office, 'name', 'Country Office')
    _title(ws, f'{co_name} — Progress Report', span=10)
    _subtitle(ws, f'Reporting status for {year} {quarter} · '
                  f'Generated {date.today():%d %b %Y}', span=10)

    # KPI strip
    summary = {'submitted': 0, 'under_review': 0, 'pending': 0,
               'overdue': 0, 'not_started': 0, 'not_applicable': 0, 'closed': 0}
    today = date.today()
    for p in projects:
        s = statuses_by_project.get(p.pk)
        code = s.status if s else 'not_started'
        if code in ('pending', 'not_started') and cycle and cycle.submission_deadline \
                and cycle.submission_deadline < today:
            code = 'overdue'
        if code in summary:
            summary[code] += 1

    total = len(list(projects))
    _section(ws, 'Summary', row=4, span=10)
    tiles = [
        ('Projects', total, DODO_BLUE),
        ('Submitted', summary['submitted'], '16A34A'),
        ('Under Review', summary['under_review'], '3B82F6'),
        ('Pending', summary['pending'], 'F59E0B'),
        ('Overdue', summary['overdue'], 'DC2626'),
    ]
    col = 1
    for label, value, color in tiles:
        _kpi_tile(ws, 6, col, label, value, color=color)
        col += 2

    # Cycle metadata
    _section(ws, 'Cycle details', row=9, span=10)
    meta = [
        ('Year', year),
        ('Quarter', quarter),
        ('Cycle configured', 'Yes' if cycle else 'No'),
        ('Submission deadline', cycle.submission_deadline.strftime('%d %b %Y')
            if cycle and cycle.submission_deadline else '—'),
        ('Final report due', cycle.final_report_due.strftime('%d %b %Y')
            if cycle and cycle.final_report_due else '—'),
        ('Reporting timeline', _val(cycle, 'reporting_timeline', '—')),
    ]
    for i, (label, value) in enumerate(meta):
        _body_cell(ws, label, 11 + i, 1, bold=True, fill=GREY_BG)
        _body_cell(ws, value, 11 + i, 2)
        ws.merge_cells(start_row=11 + i, start_column=2,
                       end_row=11 + i, end_column=10)

    _set_widths(ws, {
        'A': 24, 'B': 20, 'C': 16, 'D': 16, 'E': 16,
        'F': 16, 'G': 16, 'H': 16, 'I': 16, 'J': 16,
    })

    # ---- Status Analysis sheet (cluster × status matrix) ----
    ws_an = wb.create_sheet('Status_Analysis')
    _title(ws_an, 'Status by Programme Unit', span=9)
    _subtitle(ws_an, 'Counts of each report status per programme cluster', span=9)

    by_unit = defaultdict(lambda: defaultdict(int))
    unit_names = {}
    for p in projects:
        s = statuses_by_project.get(p.pk)
        code = s.status if s else 'not_started'
        if code in ('pending', 'not_started') and cycle and cycle.submission_deadline \
                and cycle.submission_deadline < today:
            code = 'overdue'
        unit = p.programme_unit
        unit_id = unit.pk if unit else None
        unit_names[unit_id] = unit.name if unit else 'Unassigned'
        by_unit[unit_id][code] += 1
        by_unit[unit_id]['_total'] += 1

    headers = ['Programme Unit', 'Total', 'Submitted', 'Under Review',
               'Pending', 'Overdue', 'Not Started', 'N/A', 'Completion %']
    _header_row(ws_an, headers, row=4)

    r = 5
    for uid, name in sorted(unit_names.items(), key=lambda kv: kv[1]):
        d = by_unit[uid]
        _body_cell(ws_an, name, r, 1, bold=True)
        _body_cell(ws_an, d['_total'], r, 2, align='center', bold=True)
        _body_cell(ws_an, d['submitted'], r, 3, align='center',
                   fill=STATUS_FILLS['submitted'], font_color=STATUS_FONTS['submitted'])
        _body_cell(ws_an, d['under_review'], r, 4, align='center',
                   fill=STATUS_FILLS['under_review'], font_color=STATUS_FONTS['under_review'])
        _body_cell(ws_an, d['pending'], r, 5, align='center',
                   fill=STATUS_FILLS['pending'], font_color=STATUS_FONTS['pending'])
        _body_cell(ws_an, d['overdue'], r, 6, align='center',
                   fill=STATUS_FILLS['overdue'], font_color=STATUS_FONTS['overdue'])
        _body_cell(ws_an, d['not_started'], r, 7, align='center',
                   fill=STATUS_FILLS['not_started'], font_color=STATUS_FONTS['not_started'])
        _body_cell(ws_an, d['not_applicable'], r, 8, align='center',
                   fill=STATUS_FILLS['not_applicable'], font_color=STATUS_FONTS['not_applicable'])
        # Completion as a formula so users can edit and re-flow
        _body_cell(ws_an, f'=IFERROR(C{r}/B{r}, 0)', r, 9,
                   align='center', number_format='0%')
        r += 1

    # Totals row
    if r > 5:
        _body_cell(ws_an, 'TOTAL', r, 1, bold=True, fill=GREY_BG)
        for col in range(2, 9):
            letter = get_column_letter(col)
            _body_cell(ws_an, f'=SUM({letter}5:{letter}{r-1})', r, col,
                       bold=True, fill=GREY_BG, align='center')
        _body_cell(ws_an, f'=IFERROR(C{r}/B{r}, 0)', r, 9,
                   bold=True, fill=GREY_BG, align='center', number_format='0%')

    _set_widths(ws_an, {'A': 32, 'B': 10, 'C': 14, 'D': 14, 'E': 12,
                        'F': 12, 'G': 14, 'H': 10, 'I': 14})

    # ---- All Projects detail sheet ----
    ws_all = wb.create_sheet('All_Projects')
    _title(ws_all, 'All projects — detail', span=len(PROGRESS_HEADERS))
    _subtitle(ws_all, f'Every project with its {year} {quarter} status',
              span=len(PROGRESS_HEADERS))
    _header_row(ws_all, PROGRESS_HEADERS, row=4)

    r = 5
    for p in projects:
        s = statuses_by_project.get(p.pk)
        status_code = s.status if s else 'not_started'
        if status_code in ('pending', 'not_started') and cycle and cycle.submission_deadline \
                and cycle.submission_deadline < today:
            status_code = 'overdue'
            status_label = 'Overdue'
        elif s:
            status_label = s.get_status_display()
        else:
            status_label = 'Not Started'

        _body_cell(ws_all, _val(p, 'pims_id'), r, 1)
        _body_cell(ws_all, _val(p, 'display_title') or _val(p, 'title'), r, 2, wrap=True)
        _body_cell(ws_all, _val(p.programme_unit, 'name') if p.programme_unit_id else '—', r, 3)
        _body_cell(ws_all, _val(p, 'donor', '—'), r, 4)
        _body_cell(ws_all, p.get_status_display() if hasattr(p, 'get_status_display') else '',
                   r, 5, align='center')
        _status_pill(ws_all, r, 6, status_code, status_label)
        _body_cell(ws_all, s.updated_at.strftime('%d %b %Y') if s and s.updated_at else '—',
                   r, 7, align='center')
        _body_cell(ws_all,
                   (s.updated_by.get_full_name() or s.updated_by.email)
                   if s and s.updated_by_id else '—',
                   r, 8)
        _body_cell(ws_all, (s.notes or '') if s else '', r, 9, wrap=True)
        r += 1

    _set_widths(ws_all, {'A': 14, 'B': 42, 'C': 22, 'D': 22, 'E': 18,
                         'F': 16, 'G': 14, 'H': 22, 'I': 32})
    _freeze_header(ws_all, row=5, col=3)

    # ---- Per-cluster sheets ----
    by_cluster = defaultdict(list)
    for p in projects:
        cluster = p.programme_unit.name if p.programme_unit_id else 'Unassigned'
        by_cluster[cluster].append(p)

    for cluster_name in sorted(by_cluster):
        sheet_name = _safe_sheet_name(f'{cluster_name[:24]}')
        ws_c = wb.create_sheet(sheet_name)
        _title(ws_c, f'{cluster_name} — {year} {quarter}', span=len(PROGRESS_HEADERS))
        _subtitle(ws_c, f'{len(by_cluster[cluster_name])} project(s) in this cluster',
                  span=len(PROGRESS_HEADERS))
        _header_row(ws_c, PROGRESS_HEADERS, row=4)

        r = 5
        for p in by_cluster[cluster_name]:
            s = statuses_by_project.get(p.pk)
            status_code = s.status if s else 'not_started'
            if status_code in ('pending', 'not_started') and cycle and cycle.submission_deadline \
                    and cycle.submission_deadline < today:
                status_code = 'overdue'
                status_label = 'Overdue'
            elif s:
                status_label = s.get_status_display()
            else:
                status_label = 'Not Started'

            _body_cell(ws_c, _val(p, 'pims_id'), r, 1)
            _body_cell(ws_c, _val(p, 'display_title') or _val(p, 'title'), r, 2, wrap=True)
            _body_cell(ws_c, cluster_name, r, 3)
            _body_cell(ws_c, _val(p, 'donor', '—'), r, 4)
            _body_cell(ws_c, p.get_status_display() if hasattr(p, 'get_status_display') else '',
                       r, 5, align='center')
            _status_pill(ws_c, r, 6, status_code, status_label)
            _body_cell(ws_c, s.updated_at.strftime('%d %b %Y') if s and s.updated_at else '—',
                       r, 7, align='center')
            _body_cell(ws_c,
                       (s.updated_by.get_full_name() or s.updated_by.email)
                       if s and s.updated_by_id else '—',
                       r, 8)
            _body_cell(ws_c, (s.notes or '') if s else '', r, 9, wrap=True)
            r += 1

        _set_widths(ws_c, {'A': 14, 'B': 42, 'C': 22, 'D': 22, 'E': 18,
                           'F': 16, 'G': 14, 'H': 22, 'I': 32})
        _freeze_header(ws_c, row=5, col=3)

    return _wb_to_bytes(wb)


# ===========================================================================
# Output verification report
# ===========================================================================

VERIFICATION_HEADERS = [
    'PIMS ID', 'Project', 'Cycle', 'Status', 'Field Verification',
    'Documentation Review', 'Validation Meeting', 'Verified By',
    'Verified At', 'Final Report Due',
]


def build_verification_workbook(*, country_office, verifications):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Dashboard'

    co_name = _val(country_office, 'name', 'Country Office')
    _title(ws, f'{co_name} — Output Verification Report',
           span=len(VERIFICATION_HEADERS))
    _subtitle(ws, f'Verification status across all projects · '
                  f'Generated {date.today():%d %b %Y}',
              span=len(VERIFICATION_HEADERS))

    verifications = list(verifications)

    # KPI tally
    completed = sum(1 for v in verifications if v.status == 'completed')
    in_progress = sum(1 for v in verifications
                      if v.status in ('field_verification', 'documentation_review',
                                       'validation_meeting'))
    pending = sum(1 for v in verifications if v.status == 'pending')
    not_app = sum(1 for v in verifications if v.status == 'not_applicable')
    applicable = max(len(verifications) - not_app, 0)
    completion_pct = (completed / applicable) if applicable else 0

    _section(ws, 'Summary', row=4, span=len(VERIFICATION_HEADERS))
    tiles = [
        ('Total', len(verifications), DODO_BLUE),
        ('Completed', completed, '16A34A'),
        ('In Progress', in_progress, '3B82F6'),
        ('Pending', pending, 'F59E0B'),
        ('Not Applicable', not_app, '64748B'),
    ]
    col = 1
    for label, value, color in tiles:
        _kpi_tile(ws, 6, col, label, value, color=color)
        col += 2

    _body_cell(ws, 'Completion rate', 9, 1, bold=True, fill=GREY_BG)
    c = _body_cell(ws, completion_pct, 9, 2, bold=True, font_color=DODO_BLUE,
                   number_format='0.0%', align='center')
    ws.merge_cells(start_row=9, start_column=2, end_row=9, end_column=4)

    # Status × Cycle Year matrix on Dashboard
    _section(ws, 'By cycle year', row=11, span=len(VERIFICATION_HEADERS))

    by_year = defaultdict(lambda: defaultdict(int))
    for v in verifications:
        if v.cycle_id:
            by_year[v.cycle.year][v.status] += 1
            by_year[v.cycle.year]['_total'] += 1

    headers = ['Year', 'Total', 'Pending', 'Field Verif.', 'Docs Review',
               'Validation', 'Completed', 'N/A']
    _header_row(ws, headers, row=13)
    r = 14
    for y in sorted(by_year):
        d = by_year[y]
        _body_cell(ws, y, r, 1, bold=True, align='center')
        _body_cell(ws, d['_total'], r, 2, bold=True, align='center')
        _body_cell(ws, d['pending'], r, 3, align='center',
                   fill=STATUS_FILLS['pending'], font_color=STATUS_FONTS['pending'])
        _body_cell(ws, d['field_verification'], r, 4, align='center',
                   fill=STATUS_FILLS['field_verification'], font_color=STATUS_FONTS['field_verification'])
        _body_cell(ws, d['documentation_review'], r, 5, align='center',
                   fill=STATUS_FILLS['documentation_review'], font_color=STATUS_FONTS['documentation_review'])
        _body_cell(ws, d['validation_meeting'], r, 6, align='center',
                   fill=STATUS_FILLS['validation_meeting'], font_color=STATUS_FONTS['validation_meeting'])
        _body_cell(ws, d['completed'], r, 7, align='center',
                   fill=STATUS_FILLS['completed'], font_color=STATUS_FONTS['completed'])
        _body_cell(ws, d['not_applicable'], r, 8, align='center',
                   fill=STATUS_FILLS['not_applicable'], font_color=STATUS_FONTS['not_applicable'])
        r += 1

    _set_widths(ws, {'A': 18, 'B': 12, 'C': 14, 'D': 14, 'E': 14,
                     'F': 14, 'G': 14, 'H': 12, 'I': 14, 'J': 16})

    # ---- All verifications detail sheet ----
    ws_all = wb.create_sheet('All_Verifications')
    _title(ws_all, 'All verifications', span=len(VERIFICATION_HEADERS))
    _subtitle(ws_all, 'Every recorded output verification',
              span=len(VERIFICATION_HEADERS))
    _header_row(ws_all, VERIFICATION_HEADERS, row=4)

    r = 5
    for v in verifications:
        _body_cell(ws_all, _val(v.project, 'pims_id'), r, 1)
        _body_cell(ws_all, _val(v.project, 'display_title') or _val(v.project, 'title'),
                   r, 2, wrap=True)
        _body_cell(ws_all, f'{v.cycle.year} {v.cycle.quarter}' if v.cycle_id else '—',
                   r, 3, align='center')
        _status_pill(ws_all, r, 4, v.status, v.get_status_display())
        _body_cell(ws_all, _val(v, 'field_verification_dates', '—'), r, 5)
        _body_cell(ws_all, _val(v, 'documentation_review_dates', '—'), r, 6)
        _body_cell(ws_all, _val(v, 'validation_meeting_dates', '—'), r, 7)
        _body_cell(ws_all,
                   (v.verified_by.get_full_name() or v.verified_by.email)
                   if v.verified_by_id else '—',
                   r, 8)
        _body_cell(ws_all,
                   v.verified_at.strftime('%d %b %Y') if v.verified_at else '—',
                   r, 9, align='center')
        _body_cell(ws_all,
                   v.final_report_due.strftime('%d %b %Y') if v.final_report_due else '—',
                   r, 10, align='center')
        r += 1

    _set_widths(ws_all, {'A': 14, 'B': 38, 'C': 12, 'D': 18, 'E': 22,
                         'F': 22, 'G': 22, 'H': 22, 'I': 14, 'J': 16})
    _freeze_header(ws_all, row=5, col=3)

    # ---- Per-cluster sheets ----
    by_cluster = defaultdict(list)
    for v in verifications:
        unit = v.project.programme_unit if v.project_id else None
        cluster_name = unit.name if unit else 'Unassigned'
        by_cluster[cluster_name].append(v)

    for cluster_name, items in sorted(by_cluster.items()):
        ws_c = wb.create_sheet(_safe_sheet_name(cluster_name[:24]))
        _title(ws_c, f'{cluster_name} — Verification', span=len(VERIFICATION_HEADERS))
        _subtitle(ws_c, f'{len(items)} verification(s) in this cluster',
                  span=len(VERIFICATION_HEADERS))
        _header_row(ws_c, VERIFICATION_HEADERS, row=4)

        r = 5
        for v in items:
            _body_cell(ws_c, _val(v.project, 'pims_id'), r, 1)
            _body_cell(ws_c, _val(v.project, 'display_title') or _val(v.project, 'title'),
                       r, 2, wrap=True)
            _body_cell(ws_c, f'{v.cycle.year} {v.cycle.quarter}' if v.cycle_id else '—',
                       r, 3, align='center')
            _status_pill(ws_c, r, 4, v.status, v.get_status_display())
            _body_cell(ws_c, _val(v, 'field_verification_dates', '—'), r, 5)
            _body_cell(ws_c, _val(v, 'documentation_review_dates', '—'), r, 6)
            _body_cell(ws_c, _val(v, 'validation_meeting_dates', '—'), r, 7)
            _body_cell(ws_c,
                       (v.verified_by.get_full_name() or v.verified_by.email)
                       if v.verified_by_id else '—',
                       r, 8)
            _body_cell(ws_c,
                       v.verified_at.strftime('%d %b %Y') if v.verified_at else '—',
                       r, 9, align='center')
            _body_cell(ws_c,
                       v.final_report_due.strftime('%d %b %Y') if v.final_report_due else '—',
                       r, 10, align='center')
            r += 1

        _set_widths(ws_c, {'A': 14, 'B': 38, 'C': 12, 'D': 18, 'E': 22,
                           'F': 22, 'G': 22, 'H': 22, 'I': 14, 'J': 16})
        _freeze_header(ws_c, row=5, col=3)

    return _wb_to_bytes(wb)


# ===========================================================================
# Indicator achievements report
# ===========================================================================

QUARTERS = ['Q1', 'Q2', 'Q3', 'Q4']


def build_indicators_workbook(*, country_office, year, indicators):
    """
    indicators: list of {'indicator': CPDIndicator, 'cells': [achievement_or_None × 4],
                          'latest': achievement_or_None}
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Dashboard'
    co_name = _val(country_office, 'name', 'Country Office')

    _title(ws, f'{co_name} — Indicator Achievements', span=10)
    _subtitle(ws, f'CPD indicator achievements for {year} · '
                  f'Generated {date.today():%d %b %Y}', span=10)

    # KPI tally
    tier_count = {'impact': 0, 'outcome': 0, 'output': 0}
    with_data = 0
    for row in indicators:
        tier = _val(row['indicator'].outcome, 'tier')
        if tier in tier_count:
            tier_count[tier] += 1
        if any(row['cells']):
            with_data += 1
    total = len(indicators)
    coverage = (with_data / total) if total else 0

    _section(ws, 'Summary', row=4, span=10)
    tiles = [
        ('Indicators', total, DODO_BLUE),
        ('Impact', tier_count['impact'], '7C3AED'),
        ('Outcome', tier_count['outcome'], DODO_BLUE),
        ('Output', tier_count['output'], '16A34A'),
        ('With Data', with_data, '16A34A'),
    ]
    col = 1
    for label, value, color in tiles:
        _kpi_tile(ws, 6, col, label, value, color=color)
        col += 2

    _body_cell(ws, 'Coverage', 9, 1, bold=True, fill=GREY_BG)
    _body_cell(ws, coverage, 9, 2, bold=True, font_color=DODO_BLUE,
               number_format='0.0%', align='center')
    ws.merge_cells(start_row=9, start_column=2, end_row=9, end_column=4)

    _set_widths(ws, {'A': 20, 'B': 14, 'C': 14, 'D': 14, 'E': 14,
                     'F': 14, 'G': 14, 'H': 14, 'I': 14, 'J': 14})

    # ---- Matrix sheet ----
    ws_m = wb.create_sheet('Matrix')
    headers = ['Tier', 'Outcome', 'Code', 'Description', 'Frequency',
               'Baseline', 'End Target',
               f'{year} Q1', f'{year} Q2', f'{year} Q3', f'{year} Q4',
               'Latest']
    _title(ws_m, f'CPD Indicator Matrix — {year}', span=len(headers))
    _subtitle(ws_m, 'Quarterly achievements against baseline and end target',
              span=len(headers))
    _header_row(ws_m, headers, row=4)

    r = 5
    for row in indicators:
        ind = row['indicator']
        outcome = ind.outcome
        tier = _val(outcome, 'tier', '')
        tier_color = {'impact': '7C3AED', 'outcome': DODO_BLUE,
                      'output': '16A34A'}.get(tier, '64748B')
        tier_fill = {'impact': 'F3E8FF', 'outcome': DODO_BLUE_LIGHT,
                     'output': 'DCFCE7'}.get(tier, GREY_BG)

        _body_cell(ws_m, outcome.get_tier_display() if hasattr(outcome, 'get_tier_display') else tier,
                   r, 1, bold=True, fill=tier_fill, font_color=tier_color, align='center')
        _body_cell(ws_m, _val(outcome, 'code'), r, 2, align='center', bold=True)
        _body_cell(ws_m, _val(ind, 'code'), r, 3, align='center')
        _body_cell(ws_m, _val(ind, 'description'), r, 4, wrap=True)
        _body_cell(ws_m, _val(ind, 'frequency'), r, 5, align='center')
        _body_cell(ws_m, _val(ind, 'baseline'), r, 6, wrap=True)
        _body_cell(ws_m, _val(ind, 'end_target'), r, 7, wrap=True,
                   bold=True, font_color=DODO_BLUE)

        for q_idx in range(4):
            cell_value = row['cells'][q_idx]
            display = _val(cell_value, 'achieved_value', '—') if cell_value else '—'
            _body_cell(ws_m, display, r, 8 + q_idx, align='center',
                       fill=DODO_BLUE_LIGHT if cell_value else None,
                       bold=bool(cell_value))

        latest = row['latest']
        if latest:
            _body_cell(ws_m,
                       f'{_val(latest, "year", "")} {_val(latest, "quarter", "")}',
                       r, 12, align='center', font_color='64748B')
        else:
            _body_cell(ws_m, '—', r, 12, align='center', font_color='CBD5E1')
        r += 1

    _set_widths(ws_m, {'A': 12, 'B': 14, 'C': 12, 'D': 50, 'E': 14,
                       'F': 24, 'G': 24, 'H': 14, 'I': 14, 'J': 14, 'K': 14,
                       'L': 12})
    _freeze_header(ws_m, row=5, col=4)

    # ---- Tier-specific sheets ----
    by_tier = defaultdict(list)
    for row in indicators:
        by_tier[_val(row['indicator'].outcome, 'tier', 'other')].append(row)

    for tier_code in ('impact', 'outcome', 'output'):
        rows = by_tier.get(tier_code, [])
        if not rows:
            continue
        ws_t = wb.create_sheet(_safe_sheet_name(tier_code.title()))
        _title(ws_t, f'{tier_code.title()} indicators — {year}',
               span=len(headers))
        _subtitle(ws_t, f'{len(rows)} indicator(s)', span=len(headers))
        _header_row(ws_t, headers, row=4)

        r = 5
        for row in rows:
            ind = row['indicator']
            outcome = ind.outcome
            _body_cell(ws_t, outcome.get_tier_display() if hasattr(outcome, 'get_tier_display') else tier_code,
                       r, 1, bold=True, align='center')
            _body_cell(ws_t, _val(outcome, 'code'), r, 2, align='center', bold=True)
            _body_cell(ws_t, _val(ind, 'code'), r, 3, align='center')
            _body_cell(ws_t, _val(ind, 'description'), r, 4, wrap=True)
            _body_cell(ws_t, _val(ind, 'frequency'), r, 5, align='center')
            _body_cell(ws_t, _val(ind, 'baseline'), r, 6, wrap=True)
            _body_cell(ws_t, _val(ind, 'end_target'), r, 7, wrap=True,
                       bold=True, font_color=DODO_BLUE)
            for q_idx in range(4):
                cell_value = row['cells'][q_idx]
                display = _val(cell_value, 'achieved_value', '—') if cell_value else '—'
                _body_cell(ws_t, display, r, 8 + q_idx, align='center',
                           fill=DODO_BLUE_LIGHT if cell_value else None,
                           bold=bool(cell_value))
            latest = row['latest']
            _body_cell(ws_t,
                       f'{_val(latest, "year", "")} {_val(latest, "quarter", "")}'
                       if latest else '—',
                       r, 12, align='center')
            r += 1
        _set_widths(ws_t, {'A': 12, 'B': 14, 'C': 12, 'D': 50, 'E': 14,
                           'F': 24, 'G': 24, 'H': 14, 'I': 14, 'J': 14, 'K': 14,
                           'L': 12})
        _freeze_header(ws_t, row=5, col=4)

    return _wb_to_bytes(wb)


# ===========================================================================
# Donor reporting report
# ===========================================================================

DONOR_HEADERS = [
    'Project', 'PIMS ID', 'Cluster', 'Frequency',
    'Period 1', 'Internal Draft', 'Programme Review', 'PMSU Review', 'Submission 1',
    'Period 2', 'Internal Draft', 'Programme Review', 'PMSU Review', 'Submission 2',
    'Notes',
]


def build_donor_workbook(*, country_office, timelines):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Dashboard'

    co_name = _val(country_office, 'name', 'Country Office')
    _title(ws, f'{co_name} — Donor Reporting', span=len(DONOR_HEADERS))
    _subtitle(ws, f'Reporting timelines across the donor portfolio · '
                  f'Generated {date.today():%d %b %Y}',
              span=len(DONOR_HEADERS))

    timelines = list(timelines)
    by_donor = defaultdict(list)
    for t in timelines:
        by_donor[t.donor or 'Unspecified'].append(t)

    # KPI strip
    _section(ws, 'Summary', row=4, span=len(DONOR_HEADERS))
    unique_projects = {t.project_id for t in timelines if t.project_id}
    tiles = [
        ('Timelines', len(timelines), DODO_BLUE),
        ('Donors', len(by_donor), 'F59E0B'),
        ('Projects', len(unique_projects), '16A34A'),
    ]
    col = 1
    for label, value, color in tiles:
        _kpi_tile(ws, 6, col, label, value, color=color)
        col += 2

    # Donors-at-a-glance
    _section(ws, 'Donors at a glance', row=9, span=len(DONOR_HEADERS))
    _header_row(ws, ['Donor', 'Projects', 'Quick link'], row=11)
    r = 12
    for donor, items in sorted(by_donor.items()):
        _body_cell(ws, donor, r, 1, bold=True)
        _body_cell(ws, len(items), r, 2, align='center')
        _body_cell(ws, f'See sheet "{_safe_sheet_name(donor)}"', r, 3,
                   font_color=DODO_BLUE)
        r += 1

    _set_widths(ws, {'A': 24, 'B': 12, 'C': 32, 'D': 12, 'E': 14,
                     'F': 18, 'G': 18, 'H': 18, 'I': 16, 'J': 14,
                     'K': 18, 'L': 18, 'M': 18, 'N': 16, 'O': 30})

    # ---- All-timelines detail ----
    ws_all = wb.create_sheet('All_Timelines')
    _title(ws_all, 'All donor timelines', span=len(DONOR_HEADERS))
    _subtitle(ws_all, 'Every donor reporting timeline configured',
              span=len(DONOR_HEADERS))
    headers = ['Donor'] + DONOR_HEADERS
    _header_row(ws_all, headers, row=4)

    r = 5
    for t in timelines:
        unit = t.project.programme_unit if t.project_id else None
        _body_cell(ws_all, t.donor or '—', r, 1, bold=True)
        _body_cell(ws_all, _val(t.project, 'display_title') or _val(t.project, 'title'),
                   r, 2, wrap=True)
        _body_cell(ws_all, _val(t.project, 'pims_id'), r, 3)
        _body_cell(ws_all, _val(unit, 'name', '—'), r, 4)
        _body_cell(ws_all, _val(t, 'reporting_frequency'), r, 5,
                   bold=True, fill=DODO_BLUE_LIGHT, font_color=DODO_BLUE, align='center')
        _body_cell(ws_all, _val(t, 'period_1'), r, 6, align='center')
        _body_cell(ws_all, _val(t, 'internal_draft_1'), r, 7, align='center')
        _body_cell(ws_all, _val(t, 'programme_review_1'), r, 8, align='center')
        _body_cell(ws_all, _val(t, 'pmsu_review_1'), r, 9, align='center')
        _body_cell(ws_all, _val(t, 'final_submission_1'), r, 10,
                   bold=True, align='center', fill='FEF3C7', font_color='854D0E')
        _body_cell(ws_all, _val(t, 'period_2'), r, 11, align='center')
        _body_cell(ws_all, _val(t, 'internal_draft_2'), r, 12, align='center')
        _body_cell(ws_all, _val(t, 'programme_review_2'), r, 13, align='center')
        _body_cell(ws_all, _val(t, 'pmsu_review_2'), r, 14, align='center')
        _body_cell(ws_all, _val(t, 'final_submission_2'), r, 15,
                   bold=True, align='center', fill='FEF3C7', font_color='854D0E')
        _body_cell(ws_all, _val(t, 'notes'), r, 16, wrap=True)
        r += 1

    _set_widths(ws_all, {'A': 14, 'B': 36, 'C': 14, 'D': 22, 'E': 16,
                         'F': 14, 'G': 16, 'H': 16, 'I': 16, 'J': 16,
                         'K': 14, 'L': 16, 'M': 16, 'N': 16, 'O': 16,
                         'P': 30})
    _freeze_header(ws_all, row=5, col=3)

    # ---- Per-donor sheets ----
    for donor, items in sorted(by_donor.items()):
        ws_d = wb.create_sheet(_safe_sheet_name(donor[:28]))
        _title(ws_d, f'{donor} — Reporting Schedule', span=len(DONOR_HEADERS))
        _subtitle(ws_d, f'{len(items)} project(s) under this donor',
                  span=len(DONOR_HEADERS))
        _header_row(ws_d, DONOR_HEADERS, row=4)

        r = 5
        for t in items:
            unit = t.project.programme_unit if t.project_id else None
            _body_cell(ws_d, _val(t.project, 'display_title') or _val(t.project, 'title'),
                       r, 1, wrap=True)
            _body_cell(ws_d, _val(t.project, 'pims_id'), r, 2)
            _body_cell(ws_d, _val(unit, 'name', '—'), r, 3)
            _body_cell(ws_d, _val(t, 'reporting_frequency'), r, 4,
                       bold=True, fill=DODO_BLUE_LIGHT, font_color=DODO_BLUE, align='center')
            _body_cell(ws_d, _val(t, 'period_1'), r, 5, align='center')
            _body_cell(ws_d, _val(t, 'internal_draft_1'), r, 6, align='center')
            _body_cell(ws_d, _val(t, 'programme_review_1'), r, 7, align='center')
            _body_cell(ws_d, _val(t, 'pmsu_review_1'), r, 8, align='center')
            _body_cell(ws_d, _val(t, 'final_submission_1'), r, 9,
                       bold=True, align='center', fill='FEF3C7', font_color='854D0E')
            _body_cell(ws_d, _val(t, 'period_2'), r, 10, align='center')
            _body_cell(ws_d, _val(t, 'internal_draft_2'), r, 11, align='center')
            _body_cell(ws_d, _val(t, 'programme_review_2'), r, 12, align='center')
            _body_cell(ws_d, _val(t, 'pmsu_review_2'), r, 13, align='center')
            _body_cell(ws_d, _val(t, 'final_submission_2'), r, 14,
                       bold=True, align='center', fill='FEF3C7', font_color='854D0E')
            _body_cell(ws_d, _val(t, 'notes'), r, 15, wrap=True)
            r += 1

        _set_widths(ws_d, {'A': 36, 'B': 14, 'C': 22, 'D': 16, 'E': 14,
                           'F': 18, 'G': 18, 'H': 18, 'I': 16, 'J': 14,
                           'K': 18, 'L': 18, 'M': 18, 'N': 16, 'O': 30})
        _freeze_header(ws_d, row=5, col=2)

    return _wb_to_bytes(wb)

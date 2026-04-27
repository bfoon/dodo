import io
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from statistics import mean, median, pstdev

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Count
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.text import slugify
from django.views import View

from .models import (
    Answer,
    Question,
    QuestionChoice,
    Survey,
    SurveyResponse,
    SurveySection,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _can_view_results(user):
    """Superusers, global admins, and (per-CO) anyone with relevant perms
    can see survey analytics."""
    if not user.is_authenticated:
        return False
    if user.is_superuser or getattr(user, 'is_global_admin', False):
        return True
    from apps.accounts.models import ModulePermission
    return ModulePermission.objects.filter(
        role__user_access__user=user,
        role__user_access__is_active=True,
        module__in=('surveys', 'reporting', 'users'),
        action__in=('approve', 'edit', 'export'),
    ).exists()


def _surveys_visible_to(user, co):
    """
    Return the queryset of surveys the user can see in the list, scoped to
    the active country office.

    Visibility rules:
      • superuser or global admin   → all surveys in the CO
      • survey owner / creator       → their own surveys (always)
      • CO admin (M&E)               → all surveys in the CO
      • members of the survey's project's programme unit → that survey
      • everyone else                → only public-facing active surveys
                                        within the CO they belong to
    """
    if not user.is_authenticated:
        return Survey.objects.none()
    if co is None:
        return Survey.objects.none()

    base = Survey.objects.filter(country_office=co)

    if user.is_superuser or getattr(user, 'is_global_admin', False):
        return base

    # CO admins / users with surveys-management perms see everything in CO.
    from apps.accounts.models import ModulePermission
    is_co_admin = ModulePermission.objects.filter(
        role__user_access__user=user,
        role__user_access__country_office=co,
        role__user_access__is_active=True,
        module__in=('surveys', 'users'),
        action__in=('edit', 'approve', 'delete'),
    ).exists()
    if is_co_admin:
        return base

    # Otherwise restrict: own surveys + surveys whose project is in a unit the
    # user belongs to. We OR these together.
    from django.db.models import Q
    user_unit_ids = set()
    try:
        # Common pattern: user has explicit unit memberships through their roles
        # OR through project assignments. We pull both defensively.
        from apps.projects.models import ProgrammeUnit
        # Anyone listed as a unit lead
        user_unit_ids.update(
            ProgrammeUnit.objects.filter(lead=user).values_list('pk', flat=True)
        )
        # Anyone with any responsibility on a project in a unit
        user_unit_ids.update(
            ProgrammeUnit.objects.filter(
                projects__responsibilities__user=user,
                projects__responsibilities__is_active=True,
            ).values_list('pk', flat=True).distinct()
        )
    except Exception:
        # If models aren't quite shaped this way, fall through to creator-only
        pass

    visibility = Q(created_by=user)
    if user_unit_ids:
        visibility |= Q(project__programme_unit_id__in=user_unit_ids)
    return base.filter(visibility).distinct()


_GEO_RE = re.compile(
    r'^\s*(-?\d{1,2}(?:\.\d+)?)\s*,\s*(-?\d{1,3}(?:\.\d+)?)\s*(?:,\s*(.+?))?\s*$'
)


def _parse_geo(text_value):
    if not text_value:
        return None
    m = _GEO_RE.match(str(text_value))
    if not m:
        return None
    try:
        lat = float(m.group(1)); lng = float(m.group(2))
    except (TypeError, ValueError):
        return None
    if not (-90 <= lat <= 90 and -180 <= lng <= 180):
        return None
    return {'lat': lat, 'lng': lng, 'label': (m.group(3) or '').strip() or None}


def _parse_iso_date(s):
    """Parse YYYY-MM-DD into a tz-aware datetime at midnight (start of day),
    or None if invalid. Used for the analytics filter querystring."""
    if not s:
        return None
    try:
        d = datetime.strptime(s.strip()[:10], '%Y-%m-%d')
    except (ValueError, AttributeError):
        return None
    return timezone.make_aware(d) if timezone.is_naive(d) else d


def _parse_matrix_value(text_value):
    """Matrix answers are stored as a JSON string mapping row_label -> col_label.
    Return a dict, or {} on any parse error so analytics never crashes."""
    if not text_value:
        return {}
    try:
        data = json.loads(text_value)
        if isinstance(data, dict):
            # Coerce keys/values to strings just in case
            return {str(k): str(v) for k, v in data.items() if v not in (None, '')}
    except (ValueError, TypeError):
        pass
    return {}


# Tiny built-in sentiment lexicon. We avoid pulling in NLTK/VADER as a
# dependency — for short open-text survey answers a small lexicon gets us
# 80% of the way there, and it's easy to extend.
_POS_WORDS = frozenset({
    'good','great','excellent','amazing','wonderful','fantastic','love','loved',
    'helpful','useful','nice','perfect','best','clear','easy','enjoyed','enjoy',
    'happy','satisfied','positive','improvement','improved','effective','smooth',
    'fast','quick','friendly','supportive','recommend','awesome','outstanding',
    'pleased','impressed','well','better','beneficial','valuable','informative',
})
_NEG_WORDS = frozenset({
    'bad','poor','terrible','awful','horrible','hate','hated','useless','slow',
    'difficult','confusing','confused','unclear','broken','buggy','annoying',
    'frustrated','frustrating','disappointed','disappointing','worst','worse',
    'unhelpful','unfriendly','rude','problem','problems','issue','issues','fail',
    'failed','crash','crashed','error','errors','wrong','negative','dislike',
    'unhappy','unsatisfied','complaint','painful','waste',
})
_NEGATIONS = frozenset({'not', "n't", 'no', 'never', 'nothing', 'nobody'})


def _score_sentiment(text):
    """Return ('positive'|'negative'|'neutral', score) for one text answer.
    Score is on roughly [-1, 1]. Handles simple negation by flipping the
    polarity of the next 1-2 words after a negation token."""
    if not text:
        return ('neutral', 0.0)
    tokens = re.findall(r"[a-zA-Z']+", text.lower())
    if not tokens:
        return ('neutral', 0.0)
    score = 0
    flip_next = 0
    for tok in tokens:
        if tok in _NEGATIONS:
            flip_next = 2
            continue
        polarity = 0
        if tok in _POS_WORDS:
            polarity = 1
        elif tok in _NEG_WORDS:
            polarity = -1
        if polarity and flip_next:
            polarity = -polarity
        score += polarity
        if flip_next:
            flip_next -= 1
    norm = score / max(1, len(tokens) ** 0.5)  # length-normalised
    if norm > 0.15:
        label = 'positive'
    elif norm < -0.15:
        label = 'negative'
    else:
        label = 'neutral'
    return (label, round(norm, 3))


# ---------------------------------------------------------------------------
# List, dashboard, create, detail
# ---------------------------------------------------------------------------

class SurveyListView(LoginRequiredMixin, View):
    def get(self, request):
        co = getattr(request, 'active_country_office', None)
        surveys = (
            _surveys_visible_to(request.user, co)
            .select_related('project')
            .prefetch_related('questions', 'responses')
        )
        status_filter = request.GET.get('status')
        if status_filter:
            surveys = surveys.filter(status=status_filter)

        # Status breakdown reflects only surveys the user can see, so the
        # count badges match the list contents.
        visible_all = _surveys_visible_to(request.user, co)
        status_breakdown = []
        for sc, sl in Survey.STATUS_CHOICES:
            status_breakdown.append((sc, sl, visible_all.filter(status=sc).count()))
        return render(request, 'surveys/list.html', {
            'surveys': surveys,
            'status_filter': status_filter,
            'status_breakdown': status_breakdown,
            'total_count': visible_all.count(),
            'type_choices': Survey.TYPE_CHOICES,
        })


class SurveyDashboardView(LoginRequiredMixin, View):
    def get(self, request):
        co = getattr(request, 'active_country_office', None)
        return render(request, 'surveys/dashboard.html', {'co': co})


class SurveyCreateView(LoginRequiredMixin, View):
    def get(self, request):
        return render(request, 'surveys/form.html', {
            'survey_types': Survey.TYPE_CHOICES,
        })

    def post(self, request):
        co = getattr(request, 'active_country_office', None)
        survey = Survey.objects.create(
            country_office=co,
            title=request.POST['title'],
            description=request.POST.get('description', ''),
            survey_type=request.POST.get('survey_type', 'quick'),
            instructions=request.POST.get('instructions', ''),
            is_anonymous=bool(request.POST.get('is_anonymous')),
            allow_multiple=bool(request.POST.get('allow_multiple')),
            created_by=request.user,
        )
        messages.success(request, f'Survey "{survey.title}" created.')
        return redirect('surveys:builder', pk=survey.pk)


class SurveyDetailView(LoginRequiredMixin, View):
    def get(self, request, pk):
        survey = get_object_or_404(Survey, pk=pk)
        required_count = survey.questions.filter(is_required=True).count()
        return render(request, 'surveys/detail.html', {
            'survey': survey, 'required_count': required_count,
        })


# ---------------------------------------------------------------------------
# Builder — FIXED
# ---------------------------------------------------------------------------

class SurveyBuilderView(LoginRequiredMixin, View):
    """
    GET  → render builder with the survey's questions.
    POST → handles three actions:
           - action=publish   : flip status to 'active'
           - action=close     : flip status to 'closed'
           - (no action)      : update survey metadata (title, type, etc.)
    """

    def get(self, request, pk):
        survey = get_object_or_404(Survey, pk=pk)
        # Bug #1 fix: actually pass the questions to the template.
        questions = (
            survey.questions
            .filter(is_active=True)
            .prefetch_related('choices')
            .order_by('order', 'pk')
        )
        return render(request, 'surveys/builder.html', {
            'survey': survey,
            'questions': questions,
            'question_types': Question.TYPE_CHOICES,
        })

    def post(self, request, pk):
        survey = get_object_or_404(Survey, pk=pk)
        action = (request.POST.get('action') or '').strip()

        # Bug #3 fix: handle publish / close actions from the header buttons.
        if action == 'publish':
            if not survey.questions.filter(is_active=True).exists():
                messages.error(
                    request,
                    "Can't publish a survey with no questions. Add at least one first."
                )
            else:
                survey.status = 'active'
                if not survey.start_date:
                    survey.start_date = timezone.now()
                survey.save()
                messages.success(request, f'Survey "{survey.title}" is now active.')
            return redirect('surveys:builder', pk=pk)

        if action == 'close':
            survey.status = 'closed'
            survey.end_date = survey.end_date or timezone.now()
            survey.save()
            messages.success(request, f'Survey "{survey.title}" closed.')
            return redirect('surveys:builder', pk=pk)

        # Default: metadata save
        for field in ('title', 'description', 'instructions', 'survey_type', 'status'):
            v = request.POST.get(field)
            if v is not None:
                setattr(survey, field, v)
        if 'is_anonymous' in request.POST:
            survey.is_anonymous = bool(request.POST.get('is_anonymous'))
        if 'allow_multiple' in request.POST:
            survey.allow_multiple = bool(request.POST.get('allow_multiple'))
        survey.save()
        messages.success(request, 'Survey saved.')
        return redirect('surveys:builder', pk=pk)


# ---------------------------------------------------------------------------
# Public response form
# ---------------------------------------------------------------------------

class SurveyResponseView(View):
    """
    Public-facing response form. Intentionally NOT login-required and renders
    on a standalone template (no app chrome) — survey URLs may be shared
    with anyone.

    States:
      • status='active' AND within window     → show the form
      • status='active' AND end_date passed   → show closed page (expired)
      • status='active' AND start_date future → show closed page (not_started)
      • any other status                       → show closed page (inactive)
    """
    def _gate(self, survey):
        """Return ('open', None) or ('closed', reason_code)."""
        now = timezone.now()
        if survey.status != 'active':
            return ('closed', 'inactive')
        if survey.start_date and now < survey.start_date:
            return ('closed', 'not_started')
        if survey.end_date and now > survey.end_date:
            return ('closed', 'expired')
        return ('open', None)

    def _closed_response(self, request, survey, reason_code):
        return render(request, 'surveys/closed.html', {
            'survey': survey,
            'reason_code': reason_code,
        }, status=410 if reason_code == 'expired' else 200)

    def get(self, request, pk):
        # Don't filter by status here — load the survey so we can show the
        # right closed-state page rather than a 404.
        survey = get_object_or_404(Survey, pk=pk)

        state, reason = self._gate(survey)
        if state == 'closed':
            return self._closed_response(request, survey, reason)

        questions = (
            survey.questions.filter(is_active=True)
                  .prefetch_related('choices')
                  .order_by('order', 'pk')
        )
        max_scale = max(
            (q.scale_max for q in questions if q.question_type in ('likert', 'rating')),
            default=10,
        )
        return render(request, 'surveys/respond.html', {
            'survey': survey,
            'questions': questions,
            'likert_range': range(1, int(max_scale) + 1),
            'required_count': sum(1 for q in questions if q.is_required),
        })

    def post(self, request, pk):
        survey = get_object_or_404(Survey, pk=pk)

        # Re-check at submit time — owner may have closed the survey while
        # the respondent was filling it in.
        state, reason = self._gate(survey)
        if state == 'closed':
            return self._closed_response(request, survey, reason)

        response = SurveyResponse.objects.create(
            survey=survey,
            respondent=request.user if request.user.is_authenticated and not survey.is_anonymous else None,
            respondent_name=request.POST.get('respondent_name', '').strip(),
            respondent_email=request.POST.get('respondent_email', '').strip(),
            ip_address=request.META.get('REMOTE_ADDR'),
        )
        for q in survey.questions.filter(is_active=True):
            field_name = f'q_{q.pk}'
            answer = Answer.objects.create(response=response, question=q)
            qtype = q.question_type

            if qtype in ('text', 'textarea'):
                answer.text_value = request.POST.get(field_name, '')

            elif qtype == 'number':
                try:
                    answer.number_value = float(request.POST.get(field_name, ''))
                except (TypeError, ValueError):
                    pass

            elif qtype == 'date':
                answer.date_value = request.POST.get(field_name) or None

            elif qtype in ('radio', 'dropdown', 'yes_no'):
                cid = request.POST.get(field_name)
                if cid and str(cid).isdigit():
                    answer.save()
                    answer.selected_choices.set([int(cid)])

            elif qtype == 'checkbox':
                cids = request.POST.getlist(field_name)
                if cids:
                    answer.save()
                    answer.selected_choices.set(
                        [int(c) for c in cids if str(c).isdigit()]
                    )

            elif qtype in ('likert', 'rating'):
                try:
                    answer.number_value = float(request.POST.get(field_name, ''))
                except (TypeError, ValueError):
                    pass

            elif qtype == 'geo':
                # Widget posts: q_{pk}_lat, q_{pk}_lng, q_{pk} (place name).
                # Compose "lat,lng" or "lat,lng,place" so the analytics map
                # parser can read it back.
                lat = (request.POST.get(f'{field_name}_lat') or '').strip()
                lng = (request.POST.get(f'{field_name}_lng') or '').strip()
                place = (request.POST.get(field_name) or '').strip()
                composed = ''
                try:
                    if lat and lng:
                        flat = float(lat); flng = float(lng)
                        if -90 <= flat <= 90 and -180 <= flng <= 180:
                            composed = f'{flat},{flng}'
                            if place:
                                composed = f'{composed},{place}'
                except ValueError:
                    pass
                answer.text_value = composed or place

            elif qtype == 'file':
                uploaded = request.FILES.get(field_name)
                if uploaded:
                    answer.file_value = uploaded

            elif qtype == 'ranking':
                ranks = {}
                for ch in q.choices.all():
                    raw = request.POST.get(f'{field_name}_rank_{ch.pk}')
                    if raw and raw.strip():
                        try:
                            ranks[str(ch.pk)] = int(raw)
                        except ValueError:
                            pass
                if ranks:
                    answer.text_value = json.dumps(ranks)

            elif qtype in ('matrix', 'grid'):
                # Matrix question: rows are stored as choices with order < 1000,
                # columns as choices with order >= 1000 (convention used by the
                # builder). For each row we read q_{pk}_row_{row_pk} which
                # holds the chosen column choice id. We persist as JSON mapping
                # row label -> column label for resilient analytics even if
                # choices are later edited or reordered.
                rows = [c for c in q.choices.all() if (c.order or 0) < 1000]
                cols = [c for c in q.choices.all() if (c.order or 0) >= 1000]
                col_by_id = {c.pk: c.text for c in cols}
                grid = {}
                for row in rows:
                    raw = request.POST.get(f'{field_name}_row_{row.pk}', '')
                    if raw and str(raw).isdigit():
                        col_text = col_by_id.get(int(raw))
                        if col_text:
                            grid[row.text] = col_text
                # Fallback: also accept a flat JSON blob (e.g. from a custom
                # widget that submits the whole grid in one field).
                if not grid:
                    raw_json = request.POST.get(field_name, '')
                    if raw_json:
                        parsed = _parse_matrix_value(raw_json)
                        if parsed:
                            grid = parsed
                if grid:
                    answer.text_value = json.dumps(grid)

            elif qtype == 'section_header':
                answer.delete()
                continue

            else:
                answer.text_value = request.POST.get(field_name, '')

            answer.save()

        response.is_complete = True
        response.completed_at = timezone.now()
        response.save()
        return render(request, 'surveys/thankyou.html', {'survey': survey})


# ---------------------------------------------------------------------------
# Question CRUD
# ---------------------------------------------------------------------------

class AddQuestionView(LoginRequiredMixin, View):
    """Bug #2 fix: parse the `choices` textarea (one per line) — not a list."""

    def post(self, request, pk):
        survey = get_object_or_404(Survey, pk=pk)
        order = (
            survey.questions.aggregate(m=Count('id'))['m'] or 0
        ) + 1

        question = Question.objects.create(
            survey=survey,
            question_type=request.POST['question_type'],
            text=request.POST.get('text', '').strip(),
            description=request.POST.get('description', ''),
            order=order,
            is_required=bool(request.POST.get('is_required')),
            scale_min=self._int(request.POST.get('scale_min'), 1),
            scale_max=self._int(request.POST.get('scale_max'), 5),
            scale_min_label=request.POST.get('scale_min_label', '').strip(),
            scale_max_label=request.POST.get('scale_max_label', '').strip(),
        )

        # Auto-create yes/no choices if needed
        if question.question_type == 'yes_no':
            QuestionChoice.objects.bulk_create([
                QuestionChoice(question=question, text='Yes', value='yes', order=0),
                QuestionChoice(question=question, text='No', value='no', order=1),
            ])

        # Parse the textarea: one choice per line.
        # Also accept the legacy form field name `choice` (multiple inputs)
        # so the view stays backward compatible.
        choice_lines = []
        if 'choices' in request.POST:
            raw = request.POST.get('choices', '')
            choice_lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]
        elif 'choice' in request.POST:
            choice_lines = [
                c.strip() for c in request.POST.getlist('choice') if c.strip()
            ]

        if question.question_type in ('radio', 'checkbox', 'dropdown') and choice_lines:
            QuestionChoice.objects.bulk_create([
                QuestionChoice(question=question, text=line, order=i)
                for i, line in enumerate(choice_lines)
            ])

        messages.success(request, 'Question added.')
        return redirect('surveys:builder', pk=pk)

    @staticmethod
    def _int(val, default):
        try:
            return int(val)
        except (TypeError, ValueError):
            return default


class ReorderQuestionsView(LoginRequiredMixin, View):
    def post(self, request, pk):
        survey = get_object_or_404(Survey, pk=pk)
        try:
            payload = json.loads(request.body or '{}')
        except json.JSONDecodeError:
            return JsonResponse({'ok': False, 'error': 'Invalid JSON'}, status=400)
        order_data = payload.get('order', [])
        for new_order, q_id in enumerate(order_data, start=1):
            Question.objects.filter(pk=q_id, survey=survey).update(order=new_order)
        return JsonResponse({'ok': True})


class DeleteQuestionView(LoginRequiredMixin, View):
    def post(self, request, pk):
        q = get_object_or_404(Question, pk=pk)
        survey_id = q.survey_id
        q.delete()
        # Re-tighten ordering so the visual numbers stay 1..N
        for i, remaining in enumerate(
            Question.objects.filter(survey_id=survey_id).order_by('order', 'pk'),
            start=1,
        ):
            if remaining.order != i:
                Question.objects.filter(pk=remaining.pk).update(order=i)
        messages.success(request, 'Question deleted.')
        return redirect('surveys:builder', pk=survey_id)


# ---------------------------------------------------------------------------
# Results — comprehensive analytics for admins (unchanged from prior turn)
# ---------------------------------------------------------------------------

class SurveyResultsView(LoginRequiredMixin, View):
    def get(self, request, pk):
        survey = get_object_or_404(Survey, pk=pk)
        if not _can_view_results(request.user):
            return HttpResponseForbidden(
                "You don't have permission to view survey results. "
                "M&E admins or superusers only."
            )

        # ------------------------------------------------------------------
        # Filters (read from querystring, applied to responses queryset)
        #   ?from=YYYY-MM-DD     start date (inclusive, on started_at)
        #   ?to=YYYY-MM-DD       end date   (inclusive, on started_at)
        #   ?completed=1         only completed responses
        #   ?respondent=<id>     restrict to one respondent (user pk)
        #   ?group_by=<qid>      cross-tab pivot question (categorical)
        # ------------------------------------------------------------------
        filters = self._parse_filters(request, survey)

        questions = list(
            survey.questions.filter(is_active=True)
                  .prefetch_related('choices', 'answers__selected_choices')
                  .order_by('order')
        )

        responses_qs = (
            survey.responses.select_related('respondent')
                  .prefetch_related('answers__selected_choices')
                  .order_by('-started_at')
        )
        if filters['from_dt']:
            responses_qs = responses_qs.filter(started_at__gte=filters['from_dt'])
        if filters['to_dt']:
            responses_qs = responses_qs.filter(started_at__lt=filters['to_dt'] + timedelta(days=1))
        if filters['completed_only']:
            responses_qs = responses_qs.filter(is_complete=True)
        if filters['respondent_id']:
            responses_qs = responses_qs.filter(respondent_id=filters['respondent_id'])
        responses = list(responses_qs)

        # Restrict each question's `.answers` view to the filtered response set
        # by passing a curated answer list down the chain. We do this by
        # re-fetching answers for these responses so the analyzers stay simple.
        filtered_response_ids = {r.pk for r in responses}

        total_responses = len(responses)
        complete_responses = sum(1 for r in responses if r.is_complete)
        completion_rate = round((complete_responses / total_responses) * 100, 1) if total_responses else 0.0
        avg_duration_seconds = self._avg_duration(responses)

        timeline_chart = self._timeline_chart(responses)
        qtype_chart = self._qtype_distribution_chart(questions)

        question_blocks = []
        geo_points = []
        for q in questions:
            block = self._analyze_question_filtered(q, filtered_response_ids)
            if block:
                question_blocks.append(block)
            if q.question_type == 'geo':
                geo_points.extend(self._collect_geo_points_filtered(q, filtered_response_ids))

        map_block = self._map_block(geo_points) if geo_points else None
        recent = responses[:25]

        # Cross-tab: pivot one question's answers by the chosen `group_by`
        # question (must be categorical with answers in the filtered set).
        crosstab_options = self._crosstab_options(questions)
        crosstab_block = None
        if filters['group_by_qid']:
            crosstab_block = self._build_crosstab(
                questions, filtered_response_ids, filters['group_by_qid']
            )

        # Respondent dropdown options (non-anonymous only)
        respondent_options = self._respondent_options(survey)

        return render(request, 'surveys/results.html', {
            'survey': survey,
            'total_responses': total_responses,
            'complete_responses': complete_responses,
            'completion_rate': completion_rate,
            'avg_duration': avg_duration_seconds,
            'questions': questions,
            'question_count': len(questions),
            'required_count': sum(1 for q in questions if q.is_required),
            'recent': recent,
            'timeline_chart_json': json.dumps(timeline_chart) if timeline_chart else 'null',
            'qtype_chart_json': json.dumps(qtype_chart) if qtype_chart else 'null',
            'question_blocks_json': json.dumps(question_blocks),
            'question_blocks': question_blocks,
            'map_block_json': json.dumps(map_block) if map_block else None,
            'has_geo': bool(map_block),
            # New context
            'filters': filters,
            'crosstab_options': crosstab_options,
            'crosstab_block': crosstab_block,
            'crosstab_block_json': json.dumps(crosstab_block) if crosstab_block else 'null',
            'respondent_options': respondent_options,
            'export_querystring': request.GET.urlencode(),
        })

    # ------------------------------------------------------------------
    # Filter helpers
    # ------------------------------------------------------------------
    @staticmethod
    def _parse_filters(request, survey):
        from_dt = _parse_iso_date(request.GET.get('from'))
        to_dt = _parse_iso_date(request.GET.get('to'))
        completed_only = request.GET.get('completed') in ('1', 'true', 'on')
        try:
            respondent_id = int(request.GET.get('respondent') or 0) or None
        except ValueError:
            respondent_id = None
        try:
            group_by_qid = int(request.GET.get('group_by') or 0) or None
        except ValueError:
            group_by_qid = None
        return {
            'from': request.GET.get('from') or '',
            'to': request.GET.get('to') or '',
            'from_dt': from_dt,
            'to_dt': to_dt,
            'completed_only': completed_only,
            'respondent_id': respondent_id,
            'group_by_qid': group_by_qid,
            'is_active': any([from_dt, to_dt, completed_only, respondent_id, group_by_qid]),
        }

    @staticmethod
    def _respondent_options(survey):
        if survey.is_anonymous:
            return []
        seen = {}
        for r in survey.responses.select_related('respondent').all():
            if r.respondent_id and r.respondent_id not in seen:
                u = r.respondent
                seen[r.respondent_id] = (
                    u.get_full_name() or u.email or f'User #{u.pk}'
                )
        return sorted(seen.items(), key=lambda kv: kv[1].lower())

    @staticmethod
    def _crosstab_options(questions):
        """Questions that can serve as the 'group by' axis: anything with a
        small fixed set of answer values."""
        out = []
        for q in questions:
            if q.question_type in ('radio', 'dropdown', 'yes_no'):
                out.append({'id': q.pk, 'text': q.text, 'order': q.order})
        return out

    @staticmethod
    def _avg_duration(responses):
        durations = []
        for r in responses:
            if r.is_complete and r.started_at and r.completed_at:
                d = (r.completed_at - r.started_at).total_seconds()
                if 0 <= d <= 60 * 60 * 24:
                    durations.append(d)
        return round(mean(durations)) if durations else 0

    @staticmethod
    def _timeline_chart(responses):
        if not responses:
            return None
        buckets = defaultdict(int)
        for r in responses:
            if r.started_at:
                buckets[r.started_at.date()] += 1
        if not buckets:
            return None
        days = sorted(buckets)
        start, end = days[0], days[-1]
        filled = []
        cursor = start
        while cursor <= end:
            filled.append((cursor, buckets.get(cursor, 0)))
            cursor += timedelta(days=1)
        return {
            'data': [{
                'type': 'scatter', 'mode': 'lines+markers',
                'x': [d.isoformat() for d, _ in filled],
                'y': [n for _, n in filled],
                'fill': 'tozeroy',
                'line': {'color': '#0077c8', 'width': 2.5},
                'marker': {'color': '#0077c8', 'size': 7},
                'name': 'Responses',
                'hovertemplate': '<b>%{x|%d %b %Y}</b><br>%{y} response(s)<extra></extra>',
            }],
            'layout': {
                'margin': {'l': 40, 'r': 16, 't': 16, 'b': 40},
                'xaxis': {'gridcolor': '#f1f5f9'},
                'yaxis': {'gridcolor': '#f1f5f9', 'rangemode': 'tozero'},
                'plot_bgcolor': '#ffffff', 'paper_bgcolor': '#ffffff',
                'showlegend': False, 'height': 240,
            },
        }

    @staticmethod
    def _qtype_distribution_chart(questions):
        if not questions:
            return None
        type_label = dict(Question.TYPE_CHOICES)
        c = Counter(q.question_type for q in questions)
        items = sorted(c.items(), key=lambda kv: -kv[1])
        palette = ['#0077c8', '#16a34a', '#7c3aed', '#f59e0b', '#dc2626',
                   '#0891b2', '#db2777', '#84cc16', '#64748b', '#ea580c']
        return {
            'data': [{
                'type': 'pie',
                'labels': [type_label.get(k, k) for k, _ in items],
                'values': [v for _, v in items],
                'hole': 0.55,
                'textinfo': 'label+percent',
                'textfont': {'size': 11},
                'marker': {'colors': palette[:len(items)]},
                'hovertemplate': '<b>%{label}</b><br>%{value} question(s)<extra></extra>',
            }],
            'layout': {
                'margin': {'l': 8, 'r': 8, 't': 8, 'b': 8},
                'paper_bgcolor': '#ffffff',
                'showlegend': False, 'height': 240,
            },
        }

    def _analyze_question(self, q):
        answers = list(q.answers.all())
        return self._analyze_with_answers(q, answers)

    def _analyze_question_filtered(self, q, response_ids):
        """Same as _analyze_question but only counts answers from responses
        in the given set (used for date / completion / respondent filtering)."""
        answers = [a for a in q.answers.all() if a.response_id in response_ids]
        return self._analyze_with_answers(q, answers)

    def _analyze_with_answers(self, q, answers):
        answered = sum(1 for a in answers if self._is_answered(a, q))
        skipped = len(answers) - answered
        base = {
            'id': q.pk, 'order': q.order, 'text': q.text,
            'type': q.question_type,
            'type_label': q.get_question_type_display(),
            'is_required': q.is_required,
            'description': q.description,
            'answered': answered, 'skipped': skipped, 'total': len(answers),
        }
        if q.question_type in ('radio', 'dropdown', 'yes_no'):
            base.update(self._categorical(q, answers, multi=False))
        elif q.question_type == 'checkbox':
            base.update(self._categorical(q, answers, multi=True))
        elif q.question_type in ('likert', 'rating'):
            base.update(self._numeric_scale(q, answers))
        elif q.question_type == 'number':
            base.update(self._number(q, answers))
        elif q.question_type == 'date':
            base.update(self._date(q, answers))
        elif q.question_type in ('text', 'textarea'):
            base.update(self._text(q, answers))
        elif q.question_type == 'geo':
            base.update(self._geo_summary(q, answers))
        elif q.question_type in ('matrix', 'grid'):
            base.update(self._matrix(q, answers))
        elif q.question_type == 'file':
            base.update({'render': 'simple_count',
                         'detail': f'{answered} file(s) uploaded'})
        elif q.question_type == 'section_header':
            return None
        else:
            base.update({'render': 'unsupported'})
        return base

    def _collect_geo_points_filtered(self, q, response_ids):
        result = []
        for a in q.answers.all():
            if a.response_id not in response_ids:
                continue
            p = _parse_geo(a.text_value)
            if p:
                p['question'] = q.text[:80]
                result.append(p)
        return result

    def _build_crosstab(self, questions, response_ids, group_by_qid):
        """Pivot every (suitable) question's answers by the chosen group_by.

        Returns:
          {
            'group_by_text': str,
            'group_labels': [...],   # the categories on x axis
            'pivots': [               # one entry per pivot-able target question
              {'id', 'text', 'type', 'chart': {...}}
            ]
          }
        """
        group_q = next((q for q in questions if q.pk == group_by_qid), None)
        if not group_q or group_q.question_type not in ('radio', 'dropdown', 'yes_no'):
            return None

        # Map response_id -> group label (single value: radio/dropdown/yes_no)
        resp_to_group = {}
        for a in group_q.answers.all():
            if a.response_id not in response_ids:
                continue
            choice = a.selected_choices.first()
            if choice:
                resp_to_group[a.response_id] = choice.text

        if not resp_to_group:
            return None

        group_labels_ordered = [c.text for c in group_q.choices.all().order_by('order')]
        # Filter to groups that actually have responses
        group_labels = [g for g in group_labels_ordered if g in set(resp_to_group.values())]
        if not group_labels:
            return None

        palette = ['#0077c8', '#16a34a', '#7c3aed', '#f59e0b', '#dc2626',
                   '#0891b2', '#db2777', '#84cc16', '#64748b', '#ea580c']

        pivots = []
        for q in questions:
            if q.pk == group_by_qid:
                continue
            if q.question_type in ('radio', 'dropdown', 'yes_no', 'checkbox'):
                pivot = self._crosstab_categorical(q, resp_to_group, group_labels, palette)
            elif q.question_type in ('likert', 'rating'):
                pivot = self._crosstab_scale(q, resp_to_group, group_labels, palette)
            else:
                continue
            if pivot:
                pivots.append(pivot)

        if not pivots:
            return None

        return {
            'group_by_id': group_by_qid,
            'group_by_text': group_q.text,
            'group_labels': group_labels,
            'pivots': pivots,
        }

    @staticmethod
    def _crosstab_categorical(q, resp_to_group, group_labels, palette):
        """Stacked bar: x = answer choice, color = group label, y = count."""
        choice_order = [c.text for c in q.choices.all().order_by('order')]
        if not choice_order:
            return None
        # group -> choice -> count
        grid = {g: Counter() for g in group_labels}
        for a in q.answers.all():
            g = resp_to_group.get(a.response_id)
            if not g or g not in grid:
                continue
            for ch in a.selected_choices.all():
                grid[g][ch.text] += 1
        # Drop choices that nobody picked
        nonzero = [c for c in choice_order
                   if any(grid[g].get(c, 0) for g in group_labels)]
        if not nonzero:
            return None
        traces = []
        for i, g in enumerate(group_labels):
            traces.append({
                'type': 'bar',
                'name': g,
                'x': nonzero,
                'y': [grid[g].get(c, 0) for c in nonzero],
                'marker': {'color': palette[i % len(palette)]},
                'hovertemplate': f'<b>{g}</b><br>%{{x}}: %{{y}}<extra></extra>',
            })
        return {
            'id': q.pk, 'text': q.text,
            'type': q.question_type, 'type_label': q.get_question_type_display(),
            'chart': {
                'data': traces,
                'layout': {
                    'barmode': 'stack',
                    'margin': {'l': 50, 'r': 16, 't': 16, 'b': 80},
                    'xaxis': {'tickangle': -25},
                    'yaxis': {'gridcolor': '#f1f5f9', 'rangemode': 'tozero', 'tickformat': 'd'},
                    'plot_bgcolor': '#ffffff', 'paper_bgcolor': '#ffffff',
                    'showlegend': True,
                    'legend': {'orientation': 'h', 'y': -0.25},
                    'height': 300,
                },
            },
        }

    @staticmethod
    def _crosstab_scale(q, resp_to_group, group_labels, palette):
        """For likert/rating: show the mean score by group."""
        # group -> list of values
        by_group = defaultdict(list)
        for a in q.answers.all():
            if a.number_value is None:
                continue
            g = resp_to_group.get(a.response_id)
            if g:
                by_group[g].append(float(a.number_value))
        if not any(by_group.values()):
            return None
        means = [round(mean(by_group[g]), 2) if by_group[g] else 0.0 for g in group_labels]
        counts = [len(by_group[g]) for g in group_labels]
        return {
            'id': q.pk, 'text': q.text,
            'type': q.question_type, 'type_label': q.get_question_type_display(),
            'chart': {
                'data': [{
                    'type': 'bar',
                    'x': group_labels,
                    'y': means,
                    'text': [f'{m} (n={n})' for m, n in zip(means, counts)],
                    'textposition': 'outside',
                    'marker': {'color': palette[:len(group_labels)]},
                    'hovertemplate': '<b>%{x}</b><br>Mean: %{y}<extra></extra>',
                }],
                'layout': {
                    'margin': {'l': 50, 'r': 16, 't': 24, 'b': 60},
                    'xaxis': {'tickangle': -25},
                    'yaxis': {
                        'gridcolor': '#f1f5f9', 'rangemode': 'tozero',
                        'range': [0, q.scale_max + 0.5],
                        'title': f'Mean (scale {q.scale_min}–{q.scale_max})',
                    },
                    'plot_bgcolor': '#ffffff', 'paper_bgcolor': '#ffffff',
                    'showlegend': False, 'height': 280,
                },
            },
        }

    @staticmethod
    def _is_answered(a, q):
        if q.question_type in ('text', 'textarea', 'geo'):
            return bool((a.text_value or '').strip())
        if q.question_type in ('matrix', 'grid'):
            return bool(_parse_matrix_value(a.text_value))
        if q.question_type == 'number' or q.question_type in ('likert', 'rating'):
            return a.number_value is not None
        if q.question_type == 'date':
            return a.date_value is not None
        if q.question_type in ('radio', 'checkbox', 'dropdown', 'yes_no'):
            return a.selected_choices.exists()
        if q.question_type == 'file':
            return bool(a.file_value)
        return False

    def _categorical(self, q, answers, multi=False):
        choices = list(q.choices.all().order_by('order'))
        labels = [c.text for c in choices]
        counts = Counter()
        for a in answers:
            for ch in a.selected_choices.all():
                counts[ch.text] += 1
        values = [counts.get(c.text, 0) for c in choices]
        total_picks = sum(values)
        order = sorted(range(len(labels)), key=lambda i: -values[i])
        labels_sorted = [labels[i] for i in order]
        values_sorted = [values[i] for i in order]
        chart = {
            'data': [{
                'type': 'bar', 'orientation': 'h',
                'x': values_sorted, 'y': labels_sorted,
                'marker': {'color': '#0077c8'},
                'hovertemplate': '<b>%{y}</b><br>%{x} response(s)<extra></extra>',
            }],
            'layout': {
                'margin': {'l': 160, 'r': 24, 't': 8, 'b': 32},
                'xaxis': {'gridcolor': '#f1f5f9', 'rangemode': 'tozero', 'tickformat': 'd'},
                'yaxis': {'automargin': True, 'autorange': 'reversed'},
                'plot_bgcolor': '#ffffff', 'paper_bgcolor': '#ffffff',
                'showlegend': False,
                'height': max(180, 32 * len(labels) + 60),
            },
        }
        top_label, top_val = (labels_sorted[0], values_sorted[0]) if values_sorted else (None, 0)
        return {
            'render': 'categorical', 'multi': multi, 'chart': chart,
            'total_picks': total_picks, 'top_label': top_label,
            'top_pct': round((top_val / total_picks * 100), 1) if total_picks else 0.0,
            'choice_count': len(choices),
        }

    def _numeric_scale(self, q, answers):
        values = [float(a.number_value) for a in answers if a.number_value is not None]
        if not values:
            return {'render': 'empty'}
        scale = list(range(int(q.scale_min), int(q.scale_max) + 1))
        counts = Counter(int(round(v)) for v in values)
        bar_values = [counts.get(s, 0) for s in scale]
        labels = [str(s) for s in scale]
        if q.scale_min_label and scale:
            labels[0] = f'{scale[0]} — {q.scale_min_label}'
        if q.scale_max_label and len(scale) > 1:
            labels[-1] = f'{scale[-1]} — {q.scale_max_label}'
        heat = ['#ef4444', '#f59e0b', '#eab308', '#84cc16', '#16a34a',
                '#16a34a', '#16a34a', '#16a34a', '#16a34a', '#16a34a']
        colors = [heat[min(i, len(heat) - 1)] for i in range(len(scale))]
        mean_val = mean(values)
        chart = {
            'data': [{
                'type': 'bar', 'x': labels, 'y': bar_values,
                'marker': {'color': colors},
                'text': bar_values, 'textposition': 'outside',
                'hovertemplate': '<b>Score %{x}</b><br>%{y} response(s)<extra></extra>',
            }],
            'layout': {
                'margin': {'l': 40, 'r': 16, 't': 24, 'b': 60},
                'xaxis': {'tickangle': 0},
                'yaxis': {'gridcolor': '#f1f5f9', 'rangemode': 'tozero', 'tickformat': 'd'},
                'plot_bgcolor': '#ffffff', 'paper_bgcolor': '#ffffff',
                'showlegend': False, 'height': 260,
                # Mean indicator line
                'shapes': [{
                    'type': 'line',
                    'x0': mean_val - q.scale_min, 'x1': mean_val - q.scale_min,
                    'xref': 'x', 'yref': 'paper', 'y0': 0, 'y1': 1,
                    'line': {'color': '#0f172a', 'width': 2, 'dash': 'dot'},
                }],
                'annotations': [{
                    'x': mean_val - q.scale_min, 'y': 1, 'xref': 'x', 'yref': 'paper',
                    'text': f'Mean {mean_val:.2f}', 'showarrow': False,
                    'yshift': 8, 'font': {'size': 10, 'color': '#0f172a'},
                }],
            },
        }
        return {
            'render': 'numeric_scale', 'chart': chart,
            'mean': round(mean_val, 2),
            'median': round(median(values), 2),
            'stdev': round(pstdev(values), 2) if len(values) > 1 else 0.0,
            'p25': round(self._percentile(values, 25), 2),
            'p75': round(self._percentile(values, 75), 2),
            'p90': round(self._percentile(values, 90), 2),
            'min': min(values), 'max': max(values),
            'count': len(values),
            'scale_label': f'{q.scale_min}–{q.scale_max}',
        }

    @staticmethod
    def _percentile(values, p):
        """Linear-interpolation percentile (matches numpy's default).
        Avoids a numpy dependency for what's otherwise a pure-stdlib module."""
        if not values:
            return 0.0
        s = sorted(values)
        if len(s) == 1:
            return float(s[0])
        k = (len(s) - 1) * (p / 100.0)
        f = int(k); c = min(f + 1, len(s) - 1)
        if f == c:
            return float(s[f])
        return s[f] + (s[c] - s[f]) * (k - f)

    def _number(self, q, answers):
        values = [a.number_value for a in answers if a.number_value is not None]
        if not values:
            return {'render': 'empty'}
        chart = {
            'data': [{
                'type': 'histogram', 'x': values,
                'marker': {'color': '#0077c8'},
                'hovertemplate': '<b>Range %{x}</b><br>%{y} response(s)<extra></extra>',
            }],
            'layout': {
                'margin': {'l': 40, 'r': 16, 't': 16, 'b': 40},
                'bargap': 0.05,
                'xaxis': {'gridcolor': '#f1f5f9'},
                'yaxis': {'gridcolor': '#f1f5f9', 'rangemode': 'tozero', 'tickformat': 'd'},
                'plot_bgcolor': '#ffffff', 'paper_bgcolor': '#ffffff',
                'showlegend': False, 'height': 240,
            },
        }
        return {
            'render': 'number', 'chart': chart,
            'mean': round(mean(values), 2),
            'median': round(median(values), 2),
            'stdev': round(pstdev(values), 2) if len(values) > 1 else 0.0,
            'p25': round(self._percentile(values, 25), 2),
            'p75': round(self._percentile(values, 75), 2),
            'p90': round(self._percentile(values, 90), 2),
            'min': round(min(values), 2),
            'max': round(max(values), 2),
            'count': len(values),
        }

    def _date(self, q, answers):
        dates = [a.date_value for a in answers if a.date_value]
        if not dates:
            return {'render': 'empty'}
        buckets = Counter(dates)
        days = sorted(buckets)
        chart = {
            'data': [{
                'type': 'bar',
                'x': [d.isoformat() for d in days],
                'y': [buckets[d] for d in days],
                'marker': {'color': '#7c3aed'},
                'hovertemplate': '<b>%{x|%d %b %Y}</b><br>%{y} response(s)<extra></extra>',
            }],
            'layout': {
                'margin': {'l': 40, 'r': 16, 't': 16, 'b': 40},
                'xaxis': {'gridcolor': '#f1f5f9', 'type': 'date'},
                'yaxis': {'gridcolor': '#f1f5f9', 'rangemode': 'tozero', 'tickformat': 'd'},
                'plot_bgcolor': '#ffffff', 'paper_bgcolor': '#ffffff',
                'showlegend': False, 'height': 240,
            },
        }
        return {
            'render': 'date', 'chart': chart,
            'min_date': days[0].isoformat(),
            'max_date': days[-1].isoformat(),
            'count': len(dates),
        }

    @staticmethod
    def _text(q, answers, top_n=15):
        texts = [(a.text_value or '').strip() for a in answers]
        texts = [t for t in texts if t]
        if not texts:
            return {'render': 'empty'}
        stop = {
            'the','a','an','and','or','but','is','are','was','were','i','you',
            'we','they','he','she','it','this','that','of','in','on','at','to',
            'for','with','as','by','be','been','being','have','has','had','do',
            'does','did','not','no','so','if','then','than','too','very','just',
            'my','me','our','us','their','them','his','her','its','all','any',
            'some','more','most','much','many','will','would','should','could',
            'can','may','might','from','about','into','out','up','down','also',
        }
        words = []
        # We keep per-text token streams so we can build bigrams without
        # crossing answer boundaries.
        token_streams = []
        for t in texts:
            stream = []
            for w in re.findall(r"[a-zA-Z']{3,}", t.lower()):
                if w not in stop:
                    words.append(w)
                    stream.append(w)
            token_streams.append(stream)

        # Word chart
        word_chart = None
        if words:
            top = Counter(words).most_common(top_n)
            word_chart = {
                'data': [{
                    'type': 'bar', 'orientation': 'h',
                    'x': [n for _, n in reversed(top)],
                    'y': [w for w, _ in reversed(top)],
                    'marker': {'color': '#0891b2'},
                    'hovertemplate': '<b>%{y}</b><br>mentioned %{x} time(s)<extra></extra>',
                }],
                'layout': {
                    'margin': {'l': 110, 'r': 24, 't': 8, 'b': 32},
                    'xaxis': {'gridcolor': '#f1f5f9', 'rangemode': 'tozero', 'tickformat': 'd'},
                    'yaxis': {'automargin': True},
                    'plot_bgcolor': '#ffffff', 'paper_bgcolor': '#ffffff',
                    'showlegend': False,
                    'height': max(160, 26 * len(top) + 40),
                },
            }

        # Bigrams ("phrases") within each answer
        bigram_counter = Counter()
        for stream in token_streams:
            for i in range(len(stream) - 1):
                bigram_counter[f'{stream[i]} {stream[i+1]}'] += 1
        # Drop bigrams that appear only once — they're noise
        meaningful_bigrams = [(b, n) for b, n in bigram_counter.most_common(20) if n >= 2][:8]
        phrase_chart = None
        if meaningful_bigrams:
            phrase_chart = {
                'data': [{
                    'type': 'bar', 'orientation': 'h',
                    'x': [n for _, n in reversed(meaningful_bigrams)],
                    'y': [b for b, _ in reversed(meaningful_bigrams)],
                    'marker': {'color': '#7c3aed'},
                    'hovertemplate': '<b>%{y}</b><br>%{x} time(s)<extra></extra>',
                }],
                'layout': {
                    'margin': {'l': 140, 'r': 24, 't': 8, 'b': 32},
                    'xaxis': {'gridcolor': '#f1f5f9', 'rangemode': 'tozero', 'tickformat': 'd'},
                    'yaxis': {'automargin': True},
                    'plot_bgcolor': '#ffffff', 'paper_bgcolor': '#ffffff',
                    'showlegend': False,
                    'height': max(140, 26 * len(meaningful_bigrams) + 30),
                },
            }

        # Sentiment
        sentiments = [_score_sentiment(t) for t in texts]
        sent_counts = Counter(label for label, _ in sentiments)
        avg_sentiment = round(mean(score for _, score in sentiments), 3) if sentiments else 0.0
        sentiment_chart = None
        if sum(sent_counts.values()) > 0:
            order = ['positive', 'neutral', 'negative']
            colors = {'positive': '#16a34a', 'neutral': '#94a3b8', 'negative': '#dc2626'}
            sentiment_chart = {
                'data': [{
                    'type': 'bar',
                    'x': [sent_counts.get(k, 0) for k in order],
                    'y': [k.capitalize() for k in order],
                    'orientation': 'h',
                    'marker': {'color': [colors[k] for k in order]},
                    'text': [sent_counts.get(k, 0) for k in order],
                    'textposition': 'outside',
                    'hovertemplate': '<b>%{y}</b><br>%{x} response(s)<extra></extra>',
                }],
                'layout': {
                    'margin': {'l': 80, 'r': 30, 't': 8, 'b': 24},
                    'xaxis': {'gridcolor': '#f1f5f9', 'rangemode': 'tozero', 'tickformat': 'd'},
                    'yaxis': {'automargin': True},
                    'plot_bgcolor': '#ffffff', 'paper_bgcolor': '#ffffff',
                    'showlegend': False, 'height': 160,
                },
            }

        return {
            'render': 'text',
            'word_chart': word_chart,
            'phrase_chart': phrase_chart,
            'sentiment_chart': sentiment_chart,
            'sentiment_summary': {
                'positive': sent_counts.get('positive', 0),
                'neutral': sent_counts.get('neutral', 0),
                'negative': sent_counts.get('negative', 0),
                'avg_score': avg_sentiment,
            },
            'samples': texts[:10],
            'count': len(texts),
            'avg_length': round(mean(len(t) for t in texts), 1),
        }

    @staticmethod
    def _geo_summary(q, answers):
        points = []
        for a in answers:
            p = _parse_geo(a.text_value)
            if p:
                points.append(p)
        return {
            'render': 'geo_summary',
            'count': len(points),
            'first_point': points[0] if points else None,
        }

    @staticmethod
    def _matrix(q, answers):
        """Matrix/grid analytics.

        Convention: row choices have order < 1000, column choices have
        order >= 1000. Each answer's text_value is a JSON object mapping
        row label -> column label.

        Output:
          - a row x column count matrix (heatmap)
          - per-row stats (most-common column, response rate)
          - column totals
        """
        choices = list(q.choices.all().order_by('order'))
        rows = [c.text for c in choices if (c.order or 0) < 1000]
        cols = [c.text for c in choices if (c.order or 0) >= 1000]

        # If the builder didn't follow the order convention, fall back to
        # discovering rows/cols from the actual data so analytics still works.
        if not rows or not cols:
            seen_rows, seen_cols = set(), set()
            for a in answers:
                grid = _parse_matrix_value(a.text_value)
                for r, c in grid.items():
                    seen_rows.add(r)
                    seen_cols.add(c)
            rows = rows or sorted(seen_rows)
            cols = cols or sorted(seen_cols)

        if not rows or not cols:
            return {'render': 'empty'}

        # row -> col -> count
        grid_counts = {r: Counter() for r in rows}
        total_picks = 0
        responded = 0
        for a in answers:
            grid = _parse_matrix_value(a.text_value)
            if not grid:
                continue
            responded += 1
            for r_label, c_label in grid.items():
                if r_label in grid_counts and c_label in cols:
                    grid_counts[r_label][c_label] += 1
                    total_picks += 1

        # Build z-matrix (counts) and a percentage matrix (per-row %)
        z_counts = [[grid_counts[r].get(c, 0) for c in cols] for r in rows]
        row_totals = [sum(row) for row in z_counts]
        z_pct = [
            [(v / row_totals[i] * 100) if row_totals[i] else 0 for v in row]
            for i, row in enumerate(z_counts)
        ]

        # Heatmap chart — annotated with both count and %
        annotations = []
        for i, r in enumerate(rows):
            for j, c in enumerate(cols):
                cnt = z_counts[i][j]
                if cnt == 0:
                    continue
                pct = z_pct[i][j]
                annotations.append({
                    'x': c, 'y': r, 'xref': 'x', 'yref': 'y',
                    'text': f'{cnt}<br>({pct:.0f}%)',
                    'showarrow': False,
                    'font': {'size': 11, 'color': '#0f172a' if pct < 60 else '#ffffff'},
                })

        chart = {
            'data': [{
                'type': 'heatmap',
                'x': cols, 'y': rows, 'z': z_pct,
                'colorscale': [[0, '#f8fafc'], [0.5, '#7dd3fc'], [1, '#0077c8']],
                'showscale': True,
                'colorbar': {'title': '% of row', 'thickness': 12, 'len': 0.8},
                'hovertemplate': '<b>%{y}</b> &rarr; <b>%{x}</b><br>%{z:.0f}% of row<extra></extra>',
                'xgap': 1, 'ygap': 1,
            }],
            'layout': {
                'margin': {'l': 160, 'r': 24, 't': 24, 'b': 80},
                'xaxis': {'side': 'top', 'tickangle': -25},
                'yaxis': {'automargin': True, 'autorange': 'reversed'},
                'plot_bgcolor': '#ffffff', 'paper_bgcolor': '#ffffff',
                'height': max(220, 38 * len(rows) + 120),
                'annotations': annotations,
            },
        }

        # Per-row top-pick summary
        row_summary = []
        for i, r in enumerate(rows):
            if row_totals[i] == 0:
                row_summary.append({'row': r, 'top_col': None, 'top_pct': 0.0, 'count': 0})
                continue
            top_idx = max(range(len(cols)), key=lambda j: z_counts[i][j])
            row_summary.append({
                'row': r,
                'top_col': cols[top_idx],
                'top_pct': round(z_pct[i][top_idx], 1),
                'count': row_totals[i],
            })

        # Column totals (which option got picked most overall)
        col_totals = [sum(z_counts[i][j] for i in range(len(rows))) for j in range(len(cols))]

        return {
            'render': 'matrix',
            'chart': chart,
            'rows_count': len(rows),
            'cols_count': len(cols),
            'responded': responded,
            'total_picks': total_picks,
            'row_summary': row_summary,
            'col_totals': list(zip(cols, col_totals)),
        }

    @staticmethod
    def _collect_geo_points(q):
        result = []
        for a in q.answers.all():
            p = _parse_geo(a.text_value)
            if p:
                p['question'] = q.text[:80]
                result.append(p)
        return result

    @staticmethod
    def _map_block(points):
        """Build a plain JSON payload for the Leaflet map in results.html.

        We previously returned a Plotly scattermapbox trace, but Plotly's
        built-in 'open-street-map' style is now blocked by OSM's tile servers
        (the volunteer-run tiles refuse requests without a proper Referer),
        and the custom-raster workaround was finicky. Leaflet renders raster
        tiles cleanly with no API key, so the view just hands it the raw
        points and lets the template do the rendering.
        """
        if not points:
            return None
        lats = [p['lat'] for p in points]
        lngs = [p['lng'] for p in points]
        center = {'lat': sum(lats) / len(lats), 'lng': sum(lngs) / len(lngs)}
        spread = max(max(lats) - min(lats), max(lngs) - min(lngs))
        if spread < 0.01:    zoom = 13
        elif spread < 0.1:   zoom = 11
        elif spread < 1:     zoom = 8
        elif spread < 10:    zoom = 5
        else:                zoom = 3
        markers = []
        for p in points:
            markers.append({
                'lat': p['lat'],
                'lng': p['lng'],
                'label': p.get('label') or 'Response location',
                'question': p.get('question', ''),
            })
        return {
            'center': center,
            'zoom': zoom,
            'markers': markers,
            'count': len(points),
        }

# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

class SurveyExportView(LoginRequiredMixin, View):
    """Download survey results as an Excel workbook with three sheets:
        1. Raw responses     — one row per respondent, one column per question
        2. Summary stats     — counts, %, averages, percentiles, sentiment
        3. Matrix breakdowns — full grid for each matrix/grid question

    Honors the same querystring filters as the results page:
      ?from=, ?to=, ?completed=1, ?respondent=<user_id>
    """

    def get(self, request, pk):
        survey = get_object_or_404(Survey, pk=pk)
        if not _can_view_results(request.user):
            return HttpResponseForbidden(
                "You don't have permission to export survey results."
            )

        # Lazy import — keep openpyxl optional. If it's missing, give a
        # friendly message rather than a stack trace.
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse(
                "Excel export requires the openpyxl library. "
                "Install it with: pip install openpyxl",
                status=501, content_type='text/plain',
            )

        # Apply the same filters the results page uses
        filters = SurveyResultsView._parse_filters(request, survey)
        questions = list(
            survey.questions.filter(is_active=True)
                  .prefetch_related('choices', 'answers__selected_choices')
                  .order_by('order')
        )
        responses_qs = (
            survey.responses.select_related('respondent')
                  .prefetch_related('answers__selected_choices')
                  .order_by('started_at')
        )
        if filters['from_dt']:
            responses_qs = responses_qs.filter(started_at__gte=filters['from_dt'])
        if filters['to_dt']:
            responses_qs = responses_qs.filter(started_at__lt=filters['to_dt'] + timedelta(days=1))
        if filters['completed_only']:
            responses_qs = responses_qs.filter(is_complete=True)
        if filters['respondent_id']:
            responses_qs = responses_qs.filter(respondent_id=filters['respondent_id'])
        responses = list(responses_qs)

        wb = Workbook()
        # --- Styling tokens ---
        header_fill = PatternFill('solid', fgColor='0077C8')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # --- Sheet 1: Raw responses ---
        ws_raw = wb.active
        ws_raw.title = 'Raw responses'
        # Skip section_headers in column list
        data_questions = [q for q in questions if q.question_type != 'section_header']
        headers = ['Response #', 'Respondent', 'Started', 'Completed', 'Status']
        for q in data_questions:
            headers.append(self._truncate(q.text, 80))
        ws_raw.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws_raw.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        for row_num, r in enumerate(responses, start=2):
            row = [
                f'#{r.pk}',
                self._respondent_label(r, survey),
                self._naive(r.started_at),
                self._naive(r.completed_at),
                'Complete' if r.is_complete else 'In progress',
            ]
            answers_by_q = {a.question_id: a for a in r.answers.all()}
            for q in data_questions:
                a = answers_by_q.get(q.pk)
                row.append(self._answer_to_cell(q, a))
            ws_raw.append(row)

        self._autosize(ws_raw, max_width=50)
        ws_raw.freeze_panes = 'F2'  # keep meta cols + header visible
        for row in ws_raw.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = left_wrap

        # --- Sheet 2: Summary stats ---
        ws_sum = wb.create_sheet('Summary')
        ws_sum.append(['Question #', 'Question', 'Type', 'Answered', 'Skipped',
                       'Top / Mean', 'Detail'])
        for col_idx in range(1, 8):
            cell = ws_sum.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        results_view = SurveyResultsView()
        filtered_ids = {r.pk for r in responses}
        for q in data_questions:
            block = results_view._analyze_question_filtered(q, filtered_ids)
            if not block:
                continue
            top, detail = self._summary_for_block(block)
            ws_sum.append([
                q.order, q.text, block.get('type_label', q.question_type),
                block.get('answered', 0), block.get('skipped', 0),
                top, detail,
            ])
        self._autosize(ws_sum, max_width=60)
        ws_sum.freeze_panes = 'A2'
        for row in ws_sum.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = left_wrap

        # --- Sheet 3: Matrix breakdowns ---
        matrix_qs = [q for q in data_questions if q.question_type in ('matrix', 'grid')]
        if matrix_qs:
            ws_mat = wb.create_sheet('Matrix breakdowns')
            current_row = 1
            for q in matrix_qs:
                block = results_view._analyze_question_filtered(q, filtered_ids)
                if not block or block.get('render') == 'empty':
                    # Still write a stub so the user sees it was a matrix Q
                    ws_mat.cell(row=current_row, column=1,
                                value=f'Q{q.order}. {q.text}').font = Font(bold=True, size=12)
                    ws_mat.cell(row=current_row + 1, column=1, value='No responses yet.')
                    current_row += 3
                    continue
                # Header row for this question
                title_cell = ws_mat.cell(row=current_row, column=1,
                                         value=f'Q{q.order}. {q.text}')
                title_cell.font = Font(bold=True, size=12)
                current_row += 1
                # Heatmap data — write as a count table
                # Re-derive rows/cols from row_summary + col_totals
                rows = [r['row'] for r in block['row_summary']]
                cols = [c for c, _ in block['col_totals']]
                # Header
                ws_mat.cell(row=current_row, column=1, value='').fill = header_fill
                for j, c in enumerate(cols, start=2):
                    cell = ws_mat.cell(row=current_row, column=j, value=c)
                    cell.fill = header_fill; cell.font = header_font; cell.alignment = center
                ws_mat.cell(row=current_row, column=len(cols) + 2, value='Row total').fill = header_fill
                ws_mat.cell(row=current_row, column=len(cols) + 2).font = header_font
                ws_mat.cell(row=current_row, column=len(cols) + 2).alignment = center
                current_row += 1
                # We need the actual count grid — recompute from raw answers
                count_grid = self._matrix_counts(q, filtered_ids, rows, cols)
                for i, r in enumerate(rows):
                    cell = ws_mat.cell(row=current_row, column=1, value=r)
                    cell.font = Font(bold=True)
                    row_total = 0
                    for j, c in enumerate(cols, start=2):
                        cnt = count_grid[r].get(c, 0)
                        ws_mat.cell(row=current_row, column=j, value=cnt).alignment = center
                        row_total += cnt
                    ws_mat.cell(row=current_row, column=len(cols) + 2, value=row_total).alignment = center
                    current_row += 1
                # Column totals row
                ws_mat.cell(row=current_row, column=1, value='Column total').font = Font(bold=True)
                for j, (c, total) in enumerate(block['col_totals'], start=2):
                    ws_mat.cell(row=current_row, column=j, value=total).alignment = center
                ws_mat.cell(row=current_row, column=len(cols) + 2,
                            value=block.get('total_picks', 0)).font = Font(bold=True)
                current_row += 3  # blank rows between matrices
            self._autosize(ws_mat, max_width=40)

        # Stream the file
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        filename_base = slugify(survey.title) or f'survey-{survey.pk}'
        filename = f'{filename_base}-results-{timezone.now():%Y%m%d-%H%M}.xlsx'
        resp = HttpResponse(
            out.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    @staticmethod
    def _truncate(s, n):
        s = (s or '').strip()
        return s if len(s) <= n else s[:n - 1] + '…'

    @staticmethod
    def _respondent_label(r, survey):
        if survey.is_anonymous:
            return 'Anonymous'
        if r.respondent:
            return r.respondent.get_full_name() or r.respondent.email
        return r.respondent_name or '—'

    @staticmethod
    def _naive(dt):
        """openpyxl can't write tz-aware datetimes — strip tzinfo."""
        if not dt:
            return ''
        if timezone.is_aware(dt):
            return timezone.localtime(dt).replace(tzinfo=None)
        return dt

    @staticmethod
    def _answer_to_cell(q, a):
        if a is None:
            return ''
        qt = q.question_type
        if qt in ('text', 'textarea'):
            return a.text_value or ''
        if qt == 'number' or qt in ('likert', 'rating'):
            return a.number_value if a.number_value is not None else ''
        if qt == 'date':
            return a.date_value or ''
        if qt in ('radio', 'dropdown', 'yes_no'):
            ch = a.selected_choices.first()
            return ch.text if ch else ''
        if qt == 'checkbox':
            return ', '.join(c.text for c in a.selected_choices.all())
        if qt == 'geo':
            return a.text_value or ''
        if qt == 'file':
            return getattr(a.file_value, 'name', '') if a.file_value else ''
        if qt in ('matrix', 'grid'):
            grid = _parse_matrix_value(a.text_value)
            return '; '.join(f'{r} → {c}' for r, c in grid.items())
        return a.text_value or ''

    @staticmethod
    def _summary_for_block(block):
        """Reduce an analytics block to (top_or_mean, detail_string) for the
        Summary sheet."""
        render = block.get('render')
        if render == 'categorical':
            top = block.get('top_label') or '—'
            return (top, f"{block.get('top_pct', 0)}% of {block.get('total_picks', 0)} picks")
        if render == 'numeric_scale':
            return (
                f"Mean {block.get('mean')}",
                f"median {block.get('median')}, σ {block.get('stdev')}, "
                f"p25 {block.get('p25')}, p75 {block.get('p75')}, "
                f"range {block.get('min')}–{block.get('max')}, n={block.get('count')}",
            )
        if render == 'number':
            return (
                f"Mean {block.get('mean')}",
                f"median {block.get('median')}, σ {block.get('stdev')}, "
                f"p25 {block.get('p25')}, p75 {block.get('p75')}, "
                f"range {block.get('min')}–{block.get('max')}, n={block.get('count')}",
            )
        if render == 'date':
            return (
                f"Earliest {block.get('min_date')}",
                f"latest {block.get('max_date')}, n={block.get('count')}",
            )
        if render == 'text':
            sent = block.get('sentiment_summary') or {}
            return (
                f"{block.get('count', 0)} replies",
                f"avg length {block.get('avg_length')}, "
                f"sentiment +{sent.get('positive', 0)} / "
                f"~{sent.get('neutral', 0)} / "
                f"-{sent.get('negative', 0)} "
                f"(score {sent.get('avg_score', 0)})",
            )
        if render == 'geo_summary':
            return (f"{block.get('count', 0)} locations", '')
        if render == 'matrix':
            return (
                f"{block.get('responded', 0)} responses",
                f"{block.get('rows_count')} rows × {block.get('cols_count')} cols, "
                f"{block.get('total_picks', 0)} cells filled",
            )
        if render == 'simple_count':
            return ('', block.get('detail', ''))
        return ('', '')

    @staticmethod
    def _matrix_counts(q, response_ids, rows, cols):
        """Recompute the count grid for the Matrix sheet."""
        grid = {r: Counter() for r in rows}
        col_set = set(cols)
        for a in q.answers.all():
            if a.response_id not in response_ids:
                continue
            data = _parse_matrix_value(a.text_value)
            for r_label, c_label in data.items():
                if r_label in grid and c_label in col_set:
                    grid[r_label][c_label] += 1
        return grid

    @staticmethod
    def _autosize(ws, max_width=60):
        """Approximate column auto-sizing — openpyxl doesn't have a real one."""
        from openpyxl.utils import get_column_letter
        widths = {}
        for row in ws.iter_rows(values_only=True):
            for idx, val in enumerate(row, start=1):
                if val is None:
                    continue
                length = len(str(val))
                widths[idx] = max(widths.get(idx, 10), min(length + 2, max_width))
        for idx, w in widths.items():
            ws.column_dimensions[get_column_letter(idx)].width = w